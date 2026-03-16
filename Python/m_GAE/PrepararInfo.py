#!/usr/bin/env python
# coding: utf-8

# In[8]:


import pandas as pd
import warnings
warnings.filterwarnings('ignore')
warnings.simplefilter('ignore')


from pathlib import Path
import datetime as dt
import numpy as np
import os

#Librería para conexión a base de datos 
import pyodbc
import json

# Importar librerías de otros scripts
import importlib
import CalcularUE
importlib.reload(CalcularUE)

# Librerías de mensajería y excel
from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox
import shutil



# In[9]:


def ObtenerDemFile(s_mainpath,s_XMinfo,file,tipofile,s_FechaIni,bPrint,NamePrintFile):
    spathFile=s_XMinfo + file

    # Leer arhivo de demandas
    if tipofile=='Mod':
        lHeader=['subarea','periodo','tipo',1,2,3,4,5,6,7]
    else:
        lHeader=['subarea','periodo',1,2,3,4,5,6,7]

    df_data=pd.read_csv(spathFile,sep=',',names=lHeader,encoding="ISO-8859-1" )
    df_data['subarea']=df_data['subarea'].astype(str)

    if tipofile=='Mod':
        df_data=df_data.drop('tipo',axis=1)

    #Leer archivo de mapeos
    spathFile=r'C:\Alejo\cops\Data\Mapeos.xlsx'
    mapping_df=pd.read_excel(spathFile,sheet_name='NomDemandas')

    # Inicializar fecha
    delta=dt.timedelta(days=1)
    d_fecha=pd.to_datetime(s_FechaIni)
    df_final=pd.DataFrame()

    for i in range(1,8):
        df_day=df_data[['subarea','periodo',i]]
        df_day=df_day.rename(columns={i:'valor'})
        if tipofile=='Mod':
            df_day = df_day.merge(mapping_df, left_on='subarea', right_on='SubareaXM_Mod', how='left')
        else:
            df_day = df_day.merge(mapping_df, left_on='subarea', right_on='SubareaXM_Ini', how='left')

        df_day['subarea']=df_day['SubareaCops']
        df_day=df_day[['subarea','periodo','valor']]
        df_Aux=df_day[df_day.subarea==df_day.at[0,'subarea']]
        df_Aux['subarea']='SubEcuador138'
        df_Aux['valor']=0
        df_day=pd.concat([df_day,df_Aux],axis=0)
        df_Aux['subarea']='SubEcuador230'
        df_day=pd.concat([df_day,df_Aux],axis=0)
        df_day['fecha']=d_fecha.strftime('%Y-%m-%d')

        df_final=pd.concat([df_final,df_day],axis=0)

        d_fecha=d_fecha+delta

    df_final=df_final[['fecha','subarea','periodo','valor']]
    df_final=df_final.sort_values(['fecha','subarea','periodo'])

    if bPrint==1:
        sPathFile=s_mainpath.joinpath('Print')
        df_final.to_csv(sPathFile.joinpath(NamePrintFile), index=False)

    return df_final


# In[10]:


# Función para asignar los días de la semana a cada fecha, si es festivo se trata como un domingo
import holidays
co_holidays = holidays.Colombia()

def typedays(row,weekday):

     if weekday==True:
          return row['fecha'].weekday()
     else:
          if row['fecha'] in co_holidays:
               return 'F'
          elif row['fecha'].weekday()==5:
               return 'S'
          elif row['fecha'].weekday()==6:
               return 'F'
          else:
               return 'O'


# In[11]:


def DFTipoDia(df_fecha,df_data,sHoja,Elemento,sColumn,df_Mmto):

    
    l_per=[x for x in range(1,25)] # Vector con los periodos de día
    # print(sHoja)
    # Definir tipo de día
    df_fecha['TipoDia']=df_fecha.apply(lambda row: typedays(row,weekday=False),axis=1)

    

    # Filtrar por el tipo de elemento
    df_data=df_data[(df_data['Elemento']==Elemento)]
    
    df_Res=pd.DataFrame()
    ban=0
    # Capturar la primera fecha, acá iría el loop para recorrer el dataframe con las fechas
    for ind in df_fecha.index:

        fecha=df_fecha.loc[ind,'fecha']
        tipodia=df_fecha.loc[ind,'TipoDia']
        df_Aux=df_data[(df_data['FechaIni']<=fecha) & (df_data['FechaFin']>=fecha)]
        df_Aux=df_Aux[(df_Aux['TipoDia']==tipodia)][l_per]
        df_Aux = df_Aux.melt(var_name='periodo', value_name='valor', ignore_index=False)
        df_Aux['periodo'] = df_Aux['periodo'].astype(int)
        df_Aux['fecha']=fecha
        df_Aux[sColumn]=Elemento

        if sHoja=='MinGen':
            df_Aux2=df_Mmto[(df_Mmto['fechaIni']<=fecha) & (df_Mmto['fechaFin']>=fecha) & (df_Mmto['unidad']==Elemento)]
            if df_Aux2.shape[0]>0:
                df_Aux['valor']=0
                ban=1
            elif ban==1:
                df_Aux.loc[(df_Aux['periodo'] >= 1) & (df_Aux['periodo'] <= 15), 'valor'] = 0
                ban=0


        df_Res=pd.concat([df_Res,df_Aux],axis=0)

    return df_Res


# In[12]:


def procesar_mmtos(year: str,mes: str,Carpeta:str,sPathAE:str):


    MmtosFile = sPathAE + f"{int(year):00d}" + "-" + f"{int(mes):02d}" +  r"\\" + Carpeta + r"\\Mangen.xlsx"

    # Verificar si el archivo existe
    if os.path.exists(MmtosFile):

        #Leer archivo de mapeos
        df_Mmtos=pd.read_excel(MmtosFile,sheet_name='Sheet1')

        # df_Mmtos=df_MmtosIni.copy()
        df_Mmtos.dropna(how='all', inplace=True)
        df_Mmtos.dropna(axis=1, how='all', inplace=True)
        df_Mmtos=df_Mmtos.reset_index(drop=True)

        # Find the row where the word "Consecutivo" is located in column 1
        header_row_index = df_Mmtos[df_Mmtos.iloc[:, 0] == 'Consecutivo'].index[0]

        # Remove all rows before the header row
        df_Mmtos = df_Mmtos.iloc[header_row_index:]

        # # Set the found row as the header
        df_Mmtos.columns = df_Mmtos.iloc[0]
        df_Mmtos = df_Mmtos[1:]
        df_Mmtos.dropna(axis=1, how='all', inplace=True)

        df_Mmtos['unidad']=df_Mmtos['Elemento'].str.replace(' ', '_')
        df_Mmtos['fechaIni'] = pd.to_datetime(df_Mmtos['Fecha inicio'], format='%d/%m/%Y %H:%M').dt.strftime('%Y-%m-%d')
        df_Mmtos['fechaFin'] = pd.to_datetime(df_Mmtos['Fecha fin'], format='%d/%m/%Y %H:%M').dt.strftime('%Y-%m-%d')
        df_Mmtos['Pini'] = pd.to_datetime(df_Mmtos['Fecha inicio'], format='%d/%m/%Y %H:%M').dt.strftime('%H').astype(int)+1
        df_Mmtos['Pfin'] = pd.to_datetime(df_Mmtos['Fecha fin'], format='%d/%m/%Y %H:%M').dt.strftime('%H').astype(int)+1

        # Agregar dos días a la fecha final de las unidades de guajira
        df_Mmtos.loc[df_Mmtos['unidad'].isin(['GUAJIRA_1', 'GUAJIRA_2']), 'fechaFin'] = (
            pd.to_datetime(df_Mmtos.loc[df_Mmtos['unidad'].isin(['GUAJIRA_1', 'GUAJIRA_2']), 'fechaFin']) + pd.Timedelta(days=2)
        ).dt.strftime('%Y-%m-%d')

        l_col=df_Mmtos.columns.tolist()
        df_Mmtos=df_Mmtos[l_col[-5:]]
        MmtosFile = sPathAE + f"{int(year):00d}" + "-" + f"{int(mes):02d}" +  r"\\" + Carpeta + r"\\MangenProc.csv"
        df_Mmtos.to_csv(MmtosFile, index=False)

        return 1

    else:
            # Crear ventana raíz oculta
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)  # ✅ Hace que los messagebox estén en primer plano
        messagebox.showerror("Archivo no encontrado", f"El archivo no existe en:\n{MmtosFile}", parent=root)
        return 0


def aplicar_mmto_simultaneo(df_UniGCM, df_Mmto, unidades):
    
    df_UniGCM = df_UniGCM.copy()
    df_Mmto = df_Mmto.copy()

    df_UniGCM['fecha'] = pd.to_datetime(df_UniGCM['fecha'])
    df_Mmto['fechaIni'] = pd.to_datetime(df_Mmto['fechaIni'])
    df_Mmto['fechaFin'] = pd.to_datetime(df_Mmto['fechaFin'])

    df_Mmto_exp = df_Mmto.loc[
        df_Mmto.index.repeat(
            (df_Mmto['fechaFin'] - df_Mmto['fechaIni']).dt.days + 1
        )
    ].copy()

    df_Mmto_exp['fecha'] = (
        df_Mmto_exp.groupby(level=0)
        .apply(lambda x: pd.date_range(
            x['fechaIni'].iloc[0],
            x['fechaFin'].iloc[0]
        ))
        .explode()
        .values
    )

    mmto_counts = (
        df_Mmto_exp[df_Mmto_exp['unidad'].isin(unidades)]
        .groupby('fecha')['unidad']
        .nunique()
    )

    fechas_interseccion = mmto_counts[mmto_counts == len(unidades)].index

    df_UniGCM.loc[
        df_UniGCM['fecha'].isin(fechas_interseccion),
        'valor'
    ] = 0

    return df_UniGCM

def GetFile(year,mes,Carpeta,FechaInicial,FechaFinal,FileName, DemFile, AccessFile, InfoTipoDia,FechaDem, Horizon, sPathAE):
    #Get main path and other folders
    s_mainpath=Path.cwd()

    # Rango de análisis
    s_FechaIni=FechaInicial
    s_FechaFin=FechaFinal

    #Camniar a formato fecha

    d_FechaIni=dt.datetime.strptime(s_FechaIni, '%d/%m/%Y').strftime('%Y-%m-%d')
    d_FechaFin=dt.datetime.strptime(s_FechaFin, '%d/%m/%Y').strftime('%Y-%m-%d')

    # Cargar archivo de demanda
    # DemFile=1

    #Generar el archivo de access
    # AccessFile=0

    # Cargar informción por tipo de día
    # InfoTipoDia=1


    # Procesar los mmtos a partir del archivo descargado del SIO
    if Horizon=='LT' or Horizon=='ST':
        bMmtosFile=1
    else:
        # bMmtosFile=1
        bMmtosFile=procesar_mmtos(year,mes,Carpeta,sPathAE)
    
    if bMmtosFile==0:
        # Crear ventana raíz oculta
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)  # ✅ Hace que los messagebox estén en primer plano
        messagebox.showerror('Estado del proceso','No se imprimió el archivo de excel', parent=root)
        return 'El proceso no terminó correctamente'



    fecha_formateada = pd.to_datetime(FechaDem, format="%d/%m/%Y").strftime("%Y-%m-%d")
    FechaDem=fecha_formateada
    yearDem=pd.to_datetime(FechaDem).year
    monthDem=pd.to_datetime(FechaDem).month
    datDem=pd.to_datetime(FechaDem).day


    # Descargar la dmenada a partir de los archivios de XM
    df_final=ObtenerDemFile(s_mainpath,s_XMinfo=rf'C:\Informacion_XM\Publico\DEMANDAS\Pronostico Oficial\{yearDem}-{monthDem:02d}\\',file=rf'PRON_AREAS{monthDem:02d}{datDem:02d}.txt'
                            ,tipofile='Mod',s_FechaIni=FechaDem,bPrint=0,NamePrintFile='DemMod.txt')
    
    FechaDem = pd.to_datetime(FechaDem) + dt.timedelta(days=-7)
    yearDem=pd.to_datetime(FechaDem).year
    monthDem=pd.to_datetime(FechaDem).month
    datDem=pd.to_datetime(FechaDem).day

    df_final_1=ObtenerDemFile(s_mainpath,s_XMinfo=rf'C:\Informacion_XM\Publico\DEMANDAS\Pronostico Oficial\{yearDem}-{monthDem:02d}\\',file=rf'PRON_AREAS{monthDem:02d}{datDem:02d}.txt'
                            ,tipofile='Mod',s_FechaIni=FechaDem,bPrint=0,NamePrintFile='DemMod.txt')
    
    FechaDem = pd.to_datetime(FechaDem) + dt.timedelta(days=-7)
    yearDem=pd.to_datetime(FechaDem).year
    monthDem=pd.to_datetime(FechaDem).month
    datDem=pd.to_datetime(FechaDem).day

    df_final_2=ObtenerDemFile(s_mainpath,s_XMinfo=rf'C:\Informacion_XM\Publico\DEMANDAS\Pronostico Oficial\{yearDem}-{monthDem:02d}\\',file=rf'PRON_AREAS{monthDem:02d}{datDem:02d}.txt'
                            ,tipofile='Mod',s_FechaIni=FechaDem,bPrint=0,NamePrintFile='DemMod.txt')


    fileDate=FileName
    # fileDate='May27_Ago27'
    # fileDate='Sep27_Dic27'
    s_mainpath=Path.cwd()
    s_mainpath = Path(__file__).resolve()
    s_mainpath=s_mainpath.parent.parent.parent

    if Horizon=='LT':
        # Load the Excel file
        file=r'Data\DatosEntradaCOPS_Base_LP.xlsx'
    else:
        file=r'Data\DatosEntradaCOPS_Base.xlsx'

    spathFileIn =s_mainpath.joinpath(file) 

    file=r'Data\DatosEntradaCOPS_Nacional_' + fileDate + '.xlsx'
    spathFileOut =s_mainpath.joinpath(file) 

    if AccessFile==1:
        # Crear una copia del archivo de Acceses
        # Ruta del archivo de Access original
        access_file = s_mainpath.joinpath(r'Modules\\OutputData', 'ResultsGurobiBase.accdb')

        # Ruta del archivo de Access copia con el nuevo nombre
        new_access_file = s_mainpath.joinpath(r'Modules\\OutputData', f'ResultsGurobi_Nacional_{FileName}.accdb')

        # Crear una copia del archivo de Access
        shutil.copy(access_file, new_access_file)

    file=r'Python\IPOEM_Data.xlsx'
    spathDataFile = s_mainpath.joinpath(file) 
    sheet_name='Elm_TipoDia'
    df_DataTD=pd.read_excel(spathDataFile, header=0,sheet_name=sheet_name)
    df_DataTD['FechaIni'] = pd.to_datetime(df_DataTD['FechaIni'])
    df_DataTD['FechaFin'] = pd.to_datetime(df_DataTD['FechaFin'])

    with pd.ExcelWriter(spathFileOut,engine='xlsxwriter') as writer:

        workbook=load_workbook(filename=spathFileIn, read_only=True)
            # Get the sheet names
        sheet_names = workbook.sheetnames
        sheet_names.sort()
        # Iterate over each sheet name
        for sheet in sheet_names:
            print(sheet)
            if sheet in ['Demanda']:
                # Read Excel file with pandas
                df_sheet = pd.read_excel(spathFileIn, sheet_name=sheet)
                df_sheet['fecha']=pd.to_datetime(df_sheet['fecha'])

                if DemFile==1:

                    # df_sheet = df_sheet.iloc[0:0]
                    # Eliminar todas las filas del dataframe
                    df_sheet = df_sheet.drop(df_sheet.index)
                    df_final['fecha']=pd.to_datetime(df_final['fecha'])

                    # Cargar la primera seamana
                    if df_final.shape[0]>0:
                        
                        df_final['fecha']=pd.to_datetime(df_final['fecha'])
                        
                        # Validar si tiene días festivos ademas del domingo, en caso afirmativo se copia la información del primer días ordinario
                        df_fecha=df_final[['fecha']]
                        df_fecha = df_fecha.drop_duplicates()
                        df_fecha=df_fecha.reset_index()
                        df_fecha['TipoDia']=df_fecha.apply(lambda row: typedays(row,weekday=False),axis=1)
                        df_fecha['DiaSem']=df_fecha.apply(lambda row: typedays(row,weekday=True),axis=1)
                        
                        for ind in df_fecha.index:
                            iDiaSem=df_fecha.loc[ind,'DiaSem']

                            if df_fecha.loc[ind,'TipoDia']=='F' and df_fecha.loc[ind,'DiaSem']!=6:
                                df_final_1['fecha']=pd.to_datetime(df_final_1['fecha'])
                                df_final_1['TipoDia']=df_final_1.apply(lambda row: typedays(row,weekday=False),axis=1)
                                df_final_1['DiaSem']=df_final_1.apply(lambda row: typedays(row,weekday=True),axis=1)

                                
                                df_Aux=df_final_1[(df_final_1['DiaSem']==iDiaSem)]
                                df_Aux=df_Aux.reset_index(drop=True)

                                if df_Aux.loc[0,'TipoDia']!='F':
                                    df_final = df_final[df_final['fecha'] != df_fecha.loc[ind,'fecha']]
                                    df_Aux['fecha']=df_fecha.loc[ind,'fecha']
                                    df_Aux=df_Aux.drop(columns=['TipoDia','DiaSem'])
                                    df_final=pd.concat([df_final,df_Aux],axis=0)

                                else:
                                    df_final_2['fecha']=pd.to_datetime(df_final_2['fecha'])
                                    df_final_2['TipoDia']=df_final_2.apply(lambda row: typedays(row,weekday=False),axis=1)
                                    df_final_2['DiaSem']=df_final_2.apply(lambda row: typedays(row,weekday=True),axis=1)

                                    df_Aux=df_final_2[(df_final_2['DiaSem']==iDiaSem)]
                                    df_Aux=df_Aux.reset_index(drop=True)

                                    if df_Aux.loc[0,'TipoDia']!='F':
                                        df_final = df_final[df_final['fecha'] != df_fecha.loc[ind,'fecha']]
                                        df_Aux['fecha']=df_fecha.loc[ind,'fecha']
                                        df_Aux=df_Aux.drop(columns=['TipoDia','DiaSem'])
                                        df_final=pd.concat([df_final,df_Aux],axis=0)
                                    else:
                                        # Create a simple alert message
                                        root = tk.Tk()
                                        root.withdraw()  # Hide the root window
                                        messagebox.showinfo("Alert", "No fue posible reemplazar la fecha " + str(df_fecha.loc[ind,'fecha']))                                         

                        df_sheet=pd.concat([df_sheet,df_final],axis=0)

                        # Guardar el datafrmae con la información de día festivo
                        df_final['TipoDia']=df_final.apply(lambda row: typedays(row,weekday=False),axis=1)
                        df_final['DiaSem']=df_final.apply(lambda row: typedays(row,weekday=True),axis=1)
                        df_DemFestivo=df_final[(df_final.TipoDia=='F') & (df_final.DiaSem==6)]

                        df_final=df_final.drop(columns=['TipoDia','DiaSem'])
                        df_DemFestivo=df_DemFestivo.drop(columns=['TipoDia','DiaSem'])
                        # Cargar las semanas restantes, en este caso 10 semanas
                        for i in range(1,21):
                            df_Aux=df_final.copy()
                            df_Aux['fecha'] = df_Aux['fecha'] + pd.Timedelta(days=i*7)
                            df_sheet=pd.concat([df_sheet,df_Aux],axis=0)

                        # Asignar festivos
                        df_fecha=df_sheet[['fecha']]
                        df_fecha['fecha']=pd.to_datetime(df_fecha['fecha'])
                        df_fecha = df_fecha.drop_duplicates()
                        df_fecha=df_fecha.reset_index()
                        df_fecha['TipoDia']=df_fecha.apply(lambda row: typedays(row,weekday=False),axis=1)
                        df_fecha['DiaSem']=df_fecha.apply(lambda row: typedays(row,weekday=True),axis=1)

                        # df_fecha.to_csv('Fechas.csv')
                        # Recorrer cada día para definir cuales son festivos
                        for ind in df_fecha.index:
                            # print(df_fecha.loc[ind,'fecha'],ind)
                            if df_fecha.loc[ind,'TipoDia']=='F' and df_fecha.loc[ind,'DiaSem']!=6:
                                dfecha=df_fecha.loc[ind,'fecha']
                                df_sheet = df_sheet[df_sheet['fecha'] != dfecha]
                                df_Aux=df_DemFestivo.copy()
                                df_Aux['fecha']=dfecha
                                df_sheet=pd.concat([df_sheet,df_Aux],axis=0)

                    else:
                        # Create a simple alert message
                        root = tk.Tk()
                        root.withdraw()  # Hide the root window
                        messagebox.showinfo("Alert", "No data available to process.")                    


                df_sheet['fecha']=pd.to_datetime(df_sheet['fecha']).astype(str)
                df_sheet=df_sheet.sort_values(by=['fecha','subarea', 'periodo'], ascending=[True, True, True])

                df_sheet.to_excel(writer, sheet_name=sheet, index=False)
                worksheet = writer.sheets[sheet]
                worksheet.set_column('A:A', 20)
                worksheet.set_column('B:B', 30)
                worksheet.set_column('C:C', 10)
                worksheet.set_column('D:D', 20)

                df_Car2=df_sheet[df_sheet['subarea'].isin(['SubAtlantico','SubBolivar','SubGCM'])]
                df_Car2=df_Car2.groupby(['fecha','periodo'],as_index=False)['valor'].sum().reset_index(drop=True)

                df_Car=df_sheet[df_sheet['subarea'].isin(['SubAtlantico','SubBolivar','SubGCM','SubCerromatoso','SubCordoba-Sucre'])]
                df_Car=df_Car.groupby(['fecha','periodo'],as_index=False)['valor'].sum().reset_index(drop=True)

                df_DemGCM=df_sheet[df_sheet['subarea'].isin(['SubGCM'])]
                df_DemGCM=df_DemGCM.groupby(['fecha','periodo'],as_index=False)['valor'].sum().reset_index(drop=True)

                # df_Car2.to_csv('DemCar2.csv')
                        
            elif sheet in ['ManmtosUnidad']:
                
                
                MmtosFile = sPathAE + f"{int(year):00d}" + "-" + f"{int(mes):02d}" +  r"\\" + Carpeta + r"\\MangenProc.csv"
                
                df_sheet = pd.read_csv(MmtosFile, sep=',',header=0, encoding="ISO-8859-1")

                df_Mmto=df_sheet.copy()
                df_Mmto['fechaIni']=pd.to_datetime(df_Mmto['fechaIni'])
                df_Mmto['fechaFin']=pd.to_datetime(df_Mmto['fechaFin'])

                df_sheet['fechaIni']=pd.to_datetime(df_sheet['fechaIni']).astype(str)
                df_sheet['fechaFin']=pd.to_datetime(df_sheet['fechaFin']).astype(str)
                df_sheet.to_excel(writer, sheet_name=sheet, index=False)
                worksheet = writer.sheets[sheet]
                worksheet.set_column('A:A', 40)
                worksheet.set_column('B:B', 20)
                worksheet.set_column('C:C', 20)
                worksheet.set_column('D:D', 10)
                worksheet.set_column('E:E', 10) 

            elif  sheet in ['LimImpArea']:
                # Lectura de la hoja
                df_sheet = pd.read_excel(spathFileIn, sheet_name=sheet)
                df_sheet['fecha']=pd.to_datetime(df_sheet['fecha'])

                if InfoTipoDia==1:
                    # Definir fechas para los que se van a adicionar
                    df_fecha=df_sheet.copy()
                    df_fecha = df_fecha[['fecha']]
                    df_fecha['fecha'] = pd.to_datetime(df_fecha['fecha'])
                    df_fecha = df_fecha.drop_duplicates()

                    # Adicionar límite caribe por tipo de día
                    df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.area=='Caribe')].index)
                    df_LimC=DFTipoDia(df_fecha,df_DataTD,'LimImpArea',Elemento='Caribe',sColumn='area',df_Mmto=None)
                    df_LimC=df_LimC[['fecha','area','periodo','valor']]
                    df_sheet=pd.concat([df_sheet,df_LimC],axis=0)
                    df_sheet=df_sheet.reset_index(drop=True)

                    # Adicionar límite caribe2 por tipo de día
                    df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.area=='Caribe2')].index)
                    df_LimC2=DFTipoDia(df_fecha,df_DataTD,'LimImpArea',Elemento='Caribe2',sColumn='area',df_Mmto=None)
                    df_LimC2=df_LimC2[['fecha','area','periodo','valor']]
                    df_sheet=pd.concat([df_sheet,df_LimC2],axis=0)
                    df_sheet=df_sheet.reset_index(drop=True)

                df_sheet['fecha']=pd.to_datetime(df_sheet['fecha']).astype(str)
                df_sheet=df_sheet.sort_values(by=['area','fecha', 'periodo'], ascending=[True, True, True])

                df_sheet.to_excel(writer, sheet_name=sheet, index=False)
                worksheet = writer.sheets[sheet]
                worksheet.set_column('A:A', 20)
                worksheet.set_column('B:B', 30)
                worksheet.set_column('C:C', 10)
                worksheet.set_column('D:D', 20) 

            elif  sheet in ['ZonaReqUE']:

                df_sheet = pd.read_excel(spathFileIn, sheet_name=sheet)
                df_sheet['fecha']=pd.to_datetime(df_sheet['fecha'])

                # Calcular unidades equivalentes
                [df_UniCar2, df_UniAtl,df_UniBol,df_UniGCM]=CalcularUE.GetUE(df_Car2,df_DemGCM)

                #Crear zona caribe 2
                df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.zona=='Z_Caribe2')].index) 
                df_UniCar2['zona']='Z_Caribe2'
                df_UniCar2=df_UniCar2[['fecha','zona','periodo','valor']]
                df_sheet=pd.concat([df_sheet,df_UniCar2],axis=0)
                df_sheet=df_sheet.reset_index(drop=True)

                # Crear zona de atlántico
                df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.zona=='Z_Atl')].index) 
                df_UniAtl['zona']='Z_Atl'
                df_UniAtl=df_UniAtl[['fecha','zona','periodo','valor']]
                df_sheet=pd.concat([df_sheet,df_UniAtl],axis=0)
                df_sheet=df_sheet.reset_index(drop=True)

                # Crear zona de Bolivar
                df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.zona=='Z_Bol')].index) 
                df_UniBol['zona']='Z_Bol'
                df_UniBol=df_UniBol[['fecha','zona','periodo','valor']]
                df_sheet=pd.concat([df_sheet,df_UniBol],axis=0)
                df_sheet=df_sheet.reset_index(drop=True)

                # Crear zona de GCM
                df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.zona=='Z_GCM')].index) 
                df_UniGCM['zona']='Z_GCM'
                df_UniGCM=df_UniGCM[['fecha','zona','periodo','valor']]

                # Ajustar la zona para que ponga cero donde las dos unidades de guajira están en mmto
                df_UniGCM = aplicar_mmto_simultaneo(
                    df_UniGCM,
                    df_Mmto,
                    ['GUAJIRA_1','GUAJIRA_2']
                )

                df_sheet=pd.concat([df_sheet,df_UniGCM],axis=0)
                df_sheet=df_sheet.reset_index(drop=True)



                # Cálculo de la reserva de Caribe 2
                df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.zona=='Z_Res_Car2')].index)

                df_ResC2=df_Car2.copy()
                df_ResC2['fecha'] = pd.to_datetime(df_ResC2['fecha'])
                df_ResC2=df_ResC2.rename(columns={'valor':'demanda'})
                df_LimC2['fecha'] = pd.to_datetime(df_LimC2['fecha'])
                df_LimC2=df_LimC2.rename(columns={'valor':'limite'})
                df_ResC2=df_ResC2.merge(df_LimC2,left_on=['fecha','periodo'],right_on=['fecha','periodo'], how='inner')[['fecha','periodo','demanda','limite']]
                # Factor condicional según periodo
                
                if Horizon=='LT1':
                    df_ResC2['factor'] = np.where(
                        ((df_ResC2['periodo'] >= 18) & (df_ResC2['periodo'] <= 21)),
                        0.25,
                        0.14
                    )    
                    df_ResC2['Z_Res_Car2']=((((df_ResC2['demanda'])-df_ResC2['limite'])/10)*df_ResC2['factor']).round(2)            
                else:
                    df_ResC2['factor'] = np.where(
                        ((df_ResC2['periodo'] >= 1) & (df_ResC2['periodo'] <= 6)) |
                        ((df_ResC2['periodo'] >= 18) & (df_ResC2['periodo'] <= 24)),
                        1.025,
                        1.05
                    )

                    df_ResC2['Z_Res_Car2']=(((df_ResC2['factor']*df_ResC2['demanda'])-df_ResC2['limite'])/10).round(2)  

                df_ResC2['Z_Res_Car2'] = df_ResC2['Z_Res_Car2'].apply(lambda x: max(0, x))
                df_ResC2=df_ResC2.rename(columns={'Z_Res_Car2':'valor'})
                df_ResC2['zona']='Z_Res_Car2'
                df_ResC2=df_ResC2[['fecha','zona','periodo','valor']]
                df_sheet=pd.concat([df_sheet,df_ResC2],axis=0)
                df_sheet=df_sheet.reset_index(drop=True)  

                # Cálculo de la reserva de Caribe
                df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.zona=='Z_Res_Car')].index)

                df_ResC=df_Car.copy()
                df_ResC['fecha'] = pd.to_datetime(df_ResC['fecha'])
                df_ResC=df_ResC.rename(columns={'valor':'demanda'})
                df_LimC['fecha'] = pd.to_datetime(df_LimC['fecha'])
                df_LimC=df_LimC.rename(columns={'valor':'limite'})
                df_ResC=df_ResC.merge(df_LimC,left_on=['fecha','periodo'],right_on=['fecha','periodo'], how='inner')[['fecha','periodo','demanda','limite']]
                # Factor condicional según periodo
                
                if Horizon=='LT1':
                    df_ResC['factor'] = np.where(
                        ((df_ResC['periodo'] >= 18) & (df_ResC['periodo'] <= 21)),
                        0.25,
                        0.14
                    )    
                    df_ResC['Z_Res_Car']=((((df_ResC['demanda'])-df_ResC['limite'])/10)*df_ResC['factor']).round(2)              
                else:
                    df_ResC['factor'] = np.where(
                        ((df_ResC['periodo'] >= 1) & (df_ResC['periodo'] <= 6)) |
                        ((df_ResC['periodo'] >= 18) & (df_ResC['periodo'] <= 24)),
                        1.025,
                        1.05
                    )

                    df_ResC['Z_Res_Car']=(((df_ResC['factor']*df_ResC['demanda'])-df_ResC['limite'])/10).round(2)  

                df_ResC['Z_Res_Car'] = df_ResC['Z_Res_Car'].apply(lambda x: max(0, x))
                df_ResC=df_ResC.rename(columns={'Z_Res_Car':'valor'})
                df_ResC['zona']='Z_Res_Car'
                df_ResC=df_ResC[['fecha','zona','periodo','valor']]
                df_sheet=pd.concat([df_sheet,df_ResC],axis=0)
                df_sheet=df_sheet.reset_index(drop=True)


                if InfoTipoDia==1:        
                    # Definir fechas para los que se van a adicionar
                    df_fecha=df_sheet.copy()
                    df_fecha = df_fecha[['fecha']]
                    df_fecha['fecha'] = pd.to_datetime(df_fecha['fecha'])
                    df_fecha = df_fecha.drop_duplicates()

                    # Adicionar UE en Bolivar por tipo de día
                    # df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.zona=='Z_Bol')].index)
                    # df_Res=DFTipoDia(df_fecha,df_DataTD,'ZonaReqUE',Elemento='Z_Bol',sColumn='zona',df_Mmto=df_Mmto)
                    # df_Res=df_Res[['fecha','zona','periodo','valor']]
                    # df_sheet=pd.concat([df_sheet,df_Res],axis=0)
                    # df_sheet=df_sheet.reset_index(drop=True)
                    
                    # df_sheet.to_csv('UE.csv')

                    # Adicionar UE en Z_GCM por tipo de día
                    # df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.zona=='Z_GCM')].index)
                    # df_Res=DFTipoDia(df_fecha,df_DataTD,'ZonaReqUE',Elemento='Z_GCM',sColumn='zona',df_Mmto=df_Mmto)
                    # df_Res=df_Res[['fecha','zona','periodo','valor']]
                    # df_sheet=pd.concat([df_sheet,df_Res],axis=0)
                    # df_sheet=df_sheet.reset_index(drop=True)

                    # Adicionar UE en Z_Res_Car2 por tipo de día
                    # df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.zona=='Z_Res_Car2')].index)
                    # df_Res=DFTipoDia(df_fecha,df_DataTD,'ZonaReqUE',Elemento='Z_Res_Car2',sColumn='zona',df_Mmto=df_Mmto)
                    # df_Res=df_Res[['fecha','zona','periodo','valor']]
                    # df_sheet=pd.concat([df_sheet,df_Res],axis=0)
                    # df_sheet=df_sheet.reset_index(drop=True)            

                    # # Adicionar UE en Z_Res_Car por tipo de día
                    # df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.zona=='Z_Res_Car')].index)
                    # df_Res=DFTipoDia(df_fecha,df_DataTD,'ZonaReqUE',Elemento='Z_Res_Car',sColumn='zona',df_Mmto=df_Mmto)
                    # df_Res=df_Res[['fecha','zona','periodo','valor']]
                    # df_sheet=pd.concat([df_sheet,df_Res],axis=0)
                    # df_sheet=df_sheet.reset_index(drop=True) 

                df_sheet['fecha']=pd.to_datetime(df_sheet['fecha']).astype(str)
                df_sheet=df_sheet.sort_values(by=['zona','fecha', 'periodo'], ascending=[True, True, True])

                df_sheet.to_excel(writer, sheet_name=sheet, index=False)
                worksheet = writer.sheets[sheet]
                worksheet.set_column('A:A', 20)
                worksheet.set_column('B:B', 30)
                worksheet.set_column('C:C', 10)
                worksheet.set_column('D:D', 20)   
            

            elif  sheet in ['MinGen']:
                # Lectura de la hoja
                df_sheet = pd.read_excel(spathFileIn, sheet_name=sheet)
                df_sheet['fecha']=pd.to_datetime(df_sheet['fecha'])

                if InfoTipoDia==1:
                    # Definir fechas para los que se van a adicionar
                    df_fecha=df_sheet.copy()
                    df_fecha = df_fecha[['fecha']]
                    df_fecha['fecha'] = pd.to_datetime(df_fecha['fecha'])
                    df_fecha = df_fecha.drop_duplicates()

                    # Adicionar generación mínima de TERMONORTE por tipo de día
                    df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.planta=='TERMONORTE')].index)
                    df_Res=DFTipoDia(df_fecha,df_DataTD,'MinGen',Elemento='TERMONORTE',sColumn='planta',df_Mmto=df_Mmto)
                    df_Res=df_Res[['fecha','planta','periodo','valor']]
                    df_sheet=pd.concat([df_sheet,df_Res],axis=0)
                    df_sheet=df_sheet.reset_index(drop=True)

                    # Adicionar generación mínima de PROELECTRICA_2 por tipo de día
                    df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.planta=='PROELECTRICA_2')].index)
                    df_Res=DFTipoDia(df_fecha,df_DataTD,'MinGen',Elemento='PROELECTRICA_2',sColumn='planta',df_Mmto=df_Mmto)
                    df_Res=df_Res[['fecha','planta','periodo','valor']]
                    df_sheet=pd.concat([df_sheet,df_Res],axis=0)
                    df_sheet=df_sheet.reset_index(drop=True)

                    # Adicionar generación mínima de PROELECTRICA_1 por tipo de día
                    df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.planta=='PROELECTRICA_1')].index)
                    df_Res=DFTipoDia(df_fecha,df_DataTD,'MinGen',Elemento='PROELECTRICA_1',sColumn='planta',df_Mmto=df_Mmto)
                    df_Res=df_Res[['fecha','planta','periodo','valor']]
                    df_sheet=pd.concat([df_sheet,df_Res],axis=0)
                    df_sheet=df_sheet.reset_index(drop=True)

                    # Adicionar generación mínima de TERMOCANDELARIA_CC por tipo de día
                    df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.planta=='TERMOCANDELARIA_CC')].index)
                    df_Res=DFTipoDia(df_fecha,df_DataTD,'MinGen',Elemento='TERMOCANDELARIA_CC',sColumn='planta',df_Mmto=df_Mmto)
                    df_Res=df_Res[['fecha','planta','periodo','valor']]
                    df_sheet=pd.concat([df_sheet,df_Res],axis=0)
                    df_sheet=df_sheet.reset_index(drop=True)

                    # Adicionar generación mínima de GUAJIRA_1 por tipo de día
                    df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.planta=='GUAJIRA_1')].index)
                    df_Res=DFTipoDia(df_fecha,df_DataTD,'MinGen',Elemento='GUAJIRA_1',sColumn='planta',df_Mmto=df_Mmto)
                    df_Res=df_Res[['fecha','planta','periodo','valor']]
                    df_sheet=pd.concat([df_sheet,df_Res],axis=0)
                    df_sheet=df_sheet.reset_index(drop=True)

                    # Adicionar generación mínima de GUAJIRA_2 por tipo de día
                    df_sheet=df_sheet.drop(index=df_sheet[(df_sheet.planta=='GUAJIRA_2')].index)
                    df_Res=DFTipoDia(df_fecha,df_DataTD,'MinGen',Elemento='GUAJIRA_2',sColumn='planta',df_Mmto=df_Mmto)
                    df_Res=df_Res[['fecha','planta','periodo','valor']]
                    df_sheet=pd.concat([df_sheet,df_Res],axis=0)
                    df_sheet=df_sheet.reset_index(drop=True)

                df_sheet['fecha']=pd.to_datetime(df_sheet['fecha']).astype(str)
                df_sheet=df_sheet.sort_values(by=['planta','fecha', 'periodo'], ascending=[True, True, True])

                df_sheet.to_excel(writer, sheet_name=sheet, index=False)
                worksheet = writer.sheets[sheet]
                worksheet.set_column('A:A', 20)
                worksheet.set_column('B:B', 30)
                worksheet.set_column('C:C', 10)
                worksheet.set_column('D:D', 20)

            else:
                df_sheet = pd.read_excel(spathFileIn, sheet_name=sheet)

                if 'fecha' in df_sheet.columns:
                    df_sheet['fecha']=pd.to_datetime(df_sheet['fecha']).astype(str)
                df_sheet.to_excel(writer, sheet_name=sheet, index=False)
                worksheet = writer.sheets[sheet]
                worksheet.set_column('A:A', 20)
                worksheet.set_column('B:B', 30)
                worksheet.set_column('C:C', 10)
                worksheet.set_column('D:D', 20)  
                worksheet.set_column('E:E', 20)
                worksheet.set_column('F:F', 20)     

        workbook.close()



# In[ ]:
if 1==1:
# try:
    # Ruta del archivo
    sFile=r"Parametros.json"
    script_dir = Path.cwd()
    script_dir=script_dir.parent.parent
    sPathfile=os.path.join(script_dir,r"Modules\Utils\ArchivosAux",sFile)
    sPathfile = os.path.join(r"C:\Alejo\cops\Modules\Utils\ArchivosAux",sFile)

    # Open and load the JSON file
    with open(sPathfile,'r') as f:
        data = json.load(f)

    # Almancenar los parámetros en variables python
    year=data['Parametros']['Ano']
    mes=data['Parametros']['Mes']
    Carpeta=data['Parametros']['Carpeta']
    FechaInicial=data['Parametros']['Fecha_Inicial']
    FechaFinal=data['Parametros']['Fecha_Final']
    FileName=data['Parametros']['Nombre_Archivo']
    FechaDem=data['Parametros']['Fecha_Demanda']

    Horizon=data['Parametros']['Tipo_Ejecucion']

    # Banderas para la ejecución
    DemFile=data['BanderasFile']['CB_ArchivoDem']
    AccessFile=data['BanderasFile']['CB_ArchivoAccess']
    InfoTipoDia=data['BanderasFile']['CB_TipoDia']

    # Ruta principal
    sPathAE=data['Paths']['AEnerPath']


    # Cargar bandera de demandas
    if DemFile=='Verdadero':
        DemFile=1
    else:
        DemFile=0

    # Cargar bandera de archivo access
    if AccessFile=='Verdadero':
        AccessFile=1
    else:
        AccessFile=0

    # Cargar bandera de archivo access
    if InfoTipoDia=='Verdadero':
        InfoTipoDia=1
    else:
        InfoTipoDia=0

    sMsg=GetFile(year,mes,Carpeta,FechaInicial,FechaFinal,FileName, DemFile, AccessFile, InfoTipoDia, FechaDem, Horizon, sPathAE)

    # Crear ventana raíz oculta
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)  # ✅ Hace que los messagebox estén en primer plano
    messagebox.showinfo('Estado del proceso','Se imprimió el archivo de excel de manera correcta', parent=root)

# except:

#     root = tk.Tk()
#     root.withdraw()
#     root.attributes("-topmost", True)  # ✅ Hace que los messagebox estén en primer plano
#     messagebox.showerror('Estado del proceso','Error en el proceso, por favor validar', parent=root)   

