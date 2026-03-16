#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import warnings
warnings.filterwarnings('ignore')
warnings.simplefilter('ignore')


from pathlib import Path
import datetime as dt
import numpy as np
import os
from dateutil.relativedelta import relativedelta

# Librerias para ejecutar el .exe de gams 
import subprocess

#Librerías para gurobi
import sys
import gurobipy as gp
from gurobipy import GRB

#Librería para conexión a base de datos 
import pandas as pd
import pyodbc
pyodbc.pooling = False
import json

import tkinter as tk
from tkinter import messagebox
import shutil


# In[2]:


def GetVariable(df_data,Var,Index):
    df_Var=df_data[(df_data.Variable==Var)]
    df_Var["Index"] = df_Var["Index"].str.replace(r'#.*', '', regex=True)
    df_Var[Index]=df_Var['Index'].str.split('.', expand=True)
    df_Var.drop('Index', axis=1, inplace=True)
    Index.append('valor')
    df_Var=df_Var[Index]
    # df_Var = df_Var.rename(columns={'Value': 'valor'})

    return df_Var


# In[3]:


def ConexionBD(db_path):
    # Set up the connection string
    conn_str = (
        r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
        r'DBQ=' + db_path + ';'
    )

    # Establish a connection to the database
    conn = pyodbc.connect(conn_str, autocommit=True)
    cursor = conn.cursor()

    return cursor,conn


# In[4]:


def DesconexionDB(cursor,conn):
    # Close the cursor and connection
    cursor.close()
    conn.close()


# In[5]:


def ExecuteDelete(cursor,conn,Tabla,fecha):
    # Step 1: Delete data from the table
    if fecha=='1900-12-01':
        delete_query = 'DELETE * FROM ' + Tabla
    else:
        delete_query = 'DELETE * FROM ' + Tabla + " WHERE fecha=CDATE('" + fecha + "')"
    # print(delete_query)
    cursor.execute(delete_query)
    conn.commit()


# In[6]:


def ExecuteInsert(df,cursor,conn,Tabla,l_Campos,fecha):
    sCampos='('
    sVal='('
    i=1
    for col in l_Campos:

        if i<len(l_Campos):
            sCampos=sCampos + col + ','
            sVal=sVal + '?' + ','
            
        else:
            sCampos=sCampos + col + ')'
            sVal=sVal + '?' + ')'
        i=i+1
    
    insert_query = 'INSERT INTO ' + Tabla + ' ' + sCampos +  ' VALUES ' + sVal

    # print(insert_query)

    if 'fecha' not in df.columns:
        df['fecha']=fecha
    

    df=df[l_Campos]
    # Convert DataFrame to a list of tuples
    data_tuples = list(df.itertuples(index=False, name=None))

    if len(data_tuples)>0:
        # Insert the data using executemany
        cursor.executemany(insert_query, data_tuples)

        # Commit the transaction
        conn.commit()
    else:
        stop=1

    #return insert_query


# In[7]:


def ExecuteGurobi(LPfile):
    
    model = gp.read(LPfile)

    # Set parameter to Gurobi
    model.setParam(GRB.Param.TimeLimit, 1000)
    model.setParam(GRB.Param.MIPGap, 0.0001)
    model.setParam(GRB.Param.Threads, 12) 
    model.setParam(GRB.Param.Presolve, 2) 

    if model.IsMIP == 0:
        print("Model is not a MIP")
        sys.exit(0)

    model.optimize()

    return model

    # model.dispose()
    # gp.disposeDefaultEnv()

    # Obtener algunas variables
    # var=model.getVars()
    # var1=var[2]
    # print(var1)
    # print(var1.VarName)
    # print(var1.X)
    # print(var1.index)


# In[8]:


def GetModeDataGurobi(model):

    #Almacenar las variables en un dataframe
    Values = model.getAttr("X")
    Names = model.getAttr("VarName")

    df_data = pd.DataFrame({
        'VarName': Names,
        'valor': Values
    })

    # Procesar el dataframe para que quede con los indices 
    df_data[['Variable', 'Index']] = df_data['VarName'].str.split('(', expand=True)
    # Drop the original column if no longer needed
    df_data.drop('VarName', axis=1, inplace=True)
    df_data['Index'] = df_data['Index'].str.replace(')', '', regex=False)

    return df_data

    # Ejemplo de como obtener una variable del dataframe
    # df_GenPlt = GetVariable(df_data,'vGenPlt',['planta','periodo'])
    # df_GenUni = GetVariable(df_data,'vGenUni',['unidad','periodo'])
    # df_PltConf = GetVariable(df_data,'vCommitTrmCC',['planta','conf','periodo'])


# In[9]:


def GetModeDataGAMS(s_mainpath,TipoMod):

    df_data=pd.DataFrame()
    #Leer los datos de la base de datos
    file=r'Modules\OutputData\ResultsGAMS.accdb'
    db_path = str(s_mainpath.joinpath(file))

    cursor,conn=ConexionBD(db_path)

    # Descargar la información de generación de las unidades
    query = 'SELECT uni, p, value as valor FROM  pProductUni'
    df_query = pd.read_sql_query(query, conn)
    df_query['Variable']='vGenUni'
    df_query['Index']=df_query['uni'] + '.' + df_query['p']
    df_query=df_query.drop(columns=['uni', 'p'])
    df_query=df_query[['Variable','Index','valor']]
    df_data=pd.concat([df_data,df_query],axis=0)

    # Descargar la información de generación de las plantas
    query = 'SELECT plt, p, value as valor FROM  pProductPlt'
    df_query = pd.read_sql_query(query, conn)
    df_query['Variable']='vGenPlt'
    df_query['Index']=df_query['plt'] + '.' + df_query['p']
    df_query=df_query.drop(columns=['plt', 'p'])
    df_query=df_query[['Variable','Index','valor']]
    df_data=pd.concat([df_data,df_query],axis=0)

    # Descargar la información de la configuración de las plantas de ciclo combinado
    query = 'SELECT plt_t, p, value as valor FROM  pCommitTrmConf'
    df_query = pd.read_sql_query(query, conn)
    df_query['Variable']='vCommitTrmConf'
    df_query['Index']=df_query['plt_t'] + '.' + df_query['p']
    df_query=df_query.drop(columns=['plt_t', 'p'])
    df_query=df_query[['Variable','Index','valor']]
    df_data=pd.concat([df_data,df_query],axis=0)

    # Descargar la información del unit commitment de las unidades
    query = 'SELECT uni, p, value as valor FROM  pCommitmentUni'
    df_query = pd.read_sql_query(query, conn)
    df_query['Variable']='vCommitmentUni'
    df_query['Index']=df_query['uni'] + '.' + df_query['p']
    df_query=df_query.drop(columns=['uni', 'p'])
    df_query=df_query[['Variable','Index','valor']]
    df_data=pd.concat([df_data,df_query],axis=0)

    # Descargar la información de modelo 2
    query = 'SELECT plt, p, value as valor FROM  pGenTrmPltM2'
    df_query = pd.read_sql_query(query, conn)
    df_query['Variable']='vGenTrmPltM2'
    df_query['Index']=df_query['plt'] + '.' + df_query['p']
    df_query=df_query.drop(columns=['plt', 'p'])
    df_query=df_query[['Variable','Index','valor']]
    df_data=pd.concat([df_data,df_query],axis=0)

    # Descargar la información de modelo 1 de subida
    query = 'SELECT plt, p, value as valor FROM  pGenTrmPltURM1'
    df_query = pd.read_sql_query(query, conn)
    df_query['Variable']='vGenTrmPltURM1'
    df_query['Index']=df_query['plt'] + '.' + df_query['p']
    df_query=df_query.drop(columns=['plt', 'p'])
    df_query=df_query[['Variable','Index','valor']]
    df_data=pd.concat([df_data,df_query],axis=0)

    # Descargar la información de modelo 1 de bajada
    query = 'SELECT plt, p, value as valor FROM  pGenTrmPltDRM1'
    df_query = pd.read_sql_query(query, conn)
    df_query['Variable']='vGenTrmPltDRM1'
    df_query['Index']=df_query['plt'] + '.' + df_query['p']
    df_query=df_query.drop(columns=['plt', 'p'])
    df_query=df_query[['Variable','Index','valor']]
    df_data=pd.concat([df_data,df_query],axis=0)

    if TipoMod!='DESPACHO':
        # Descargar la información de precio marginal
        query = 'SELECT p, value as valor FROM  pMarginal'
        df_query = pd.read_sql_query(query, conn)
        df_query['Variable']='vMarginal'
        df_query['Index']=df_query['p']
        df_query=df_query.drop(columns=['p'])
        df_query=df_query[['Variable','Index','valor']]
        df_data=pd.concat([df_data,df_query],axis=0)

    DesconexionDB(cursor,conn)

    return df_data


# In[10]:


def PrintDataBD(df_data, pVar, pIndex, pTabla, pCampos,sfecha,pVarMod,cursor,conn):

    if pVarMod=='Mod':
        # Insertar generación de las plantas
        df_Var= GetVariable(df_data,pVar,pIndex)
        df_Var['valor']=df_Var['valor'].round(2)
    else:
        df_Var=df_data

    ExecuteDelete(cursor,conn,Tabla=pTabla,fecha=sfecha)
    ExecuteInsert(df_Var,cursor,conn,Tabla=pTabla,l_Campos=pCampos,fecha=sfecha)

    # # Insertar generación de las unidades
    # df_Var= GetVariable(df_data,'vGenUni',['unidad','periodo'])
    # ExecuteDelete(cursor,conn,Tabla='GeneracionUni',fecha=sfecha)
    # ExecuteInsert(df_Var,cursor,conn,Tabla='GeneracionUni',l_Campos=['fecha','unidad','periodo','valor'],fecha=sfecha)

    # # Insertar configuración de las plantas CC
    # df_Var= GetVariable(df_data,'vCommitTrmCC',['planta','conf','periodo'])
    # df_Var=df_Var[(df_Var.valor>0)]
    # ExecuteDelete(cursor,conn,Tabla='ConfiguracionPlt',fecha=sfecha)
    # ExecuteInsert(df_Var,cursor,conn,Tabla='ConfiguracionPlt',l_Campos=['fecha','planta','conf','periodo','valor'],fecha=sfecha)


# In[11]:


def ExecuteQuery(pTable,sFecha,conn):
    # pTable='CIUni'
    # sFecha='2024-09-04'
    # Define your query
    query = 'SELECT * FROM ' + pTable + " WHERE fecha=cdate('" + sFecha + "')"

    # Execute the query and store the result in a DataFrame
    df_table = pd.read_sql_query(query, conn)

    return df_table


# In[12]:


def printFile(df_table,fileName,index,s_mainpath):

    # Define the width for each column
    widths=[]
    for i in range(len(list(df_table.columns))):
        if i==0:
            widths.append(40)
        else:
            if fileName=='tmp_Demand.txt':
                widths.append(25)
            else:
                widths.append(15)

    if 'fecha' in df_table.columns:
        df_table=df_table.drop(['fecha'],axis=1).set_index(index)
    else:
        df_table=df_table.set_index(index)

    # Inicialización de las condiciones iniciales
    file=r'Modules\InputData\\' + fileName
    s_filepath=s_mainpath.joinpath(file)

    if fileName=='tmp_Param2.txt':
        # Open a file to write
        with open(s_filepath, 'w') as file:
            for index, row in df_table.iterrows():
                row_str = f"{index:<{widths[0]}}"  # Print index with first column width
                row_str +='     =     '
                row_str += ''.join(f'{value:<{width}}' for value, width in zip(row.values, widths[1:]))
                row_str +=';'
                file.write(row_str + '\n')
    else:
        # Open a file to write
        with open(s_filepath, 'w') as file:
            # Write the header
            header = ' ' * widths[0] + ''.join(f'{col:<{width}}' for col, width in zip(df_table.columns, widths[1:]))
            file.write(header + '\n')
            
            # Write the data rows
            for index, row in df_table.iterrows():
                row_str = f"{index:<{widths[0]}}"  # Print index with first column width
                row_str += ''.join(f'{value:<{width}}' for value, width in zip(row.values, widths[1:]))
                file.write(row_str + '\n')


# In[13]:


def CalcularCIUni(df_data,df_tableCIUni):
    df_tableCIUni=df_tableCIUni.drop('fecha',axis=1).set_index('unidad')

    df_VarGen= GetVariable(df_data,'vGenUni',['unidad','periodo'])
    df_VarGen['periodo']=df_VarGen['periodo'].astype(int)

    df_vComUni=GetVariable(df_data,'vCommitmentUni',['unidad','periodo'])
    df_vComUni['periodo']=df_vComUni['periodo'].astype(int)

    for unidad in df_tableCIUni.index:
        # unidad=df_tableCIUni.loc[index,'unidad']
        # Ajustar generación
        # if unidad=='PROELECTRICA_1':
        #     stop=1
        Gen=df_VarGen.loc[(df_VarGen.unidad==unidad)  & (df_VarGen.periodo==24),'valor'].iloc[0]
        df_tableCIUni.at[unidad, 'GenIni']=Gen

        # Ajustar tiempo en línea
        vCommitU=df_vComUni.loc[(df_vComUni.unidad==unidad)  & (df_vComUni.periodo==24),'valor'].iloc[0]
        vCommitU=round(vCommitU,0)
        if vCommitU==1:
            TLF=0
            TL=0
            for i in range(24,0,-1):
                vCommitU=df_vComUni.loc[(df_vComUni.unidad==unidad)  & (df_vComUni.periodo==i),'valor'].iloc[0]
                vCommitU=round(vCommitU,0)
                if vCommitU==1:
                    TL=TL+1
                else:
                    df_tableCIUni.at[unidad, 'TLIni']=TL
                    df_tableCIUni.at[unidad, 'TFLIni']=TLF
                    break
            
            if i==1 and vCommitU==1:
                TL=TL+df_tableCIUni.at[unidad, 'TLIni']
                df_tableCIUni.at[unidad, 'TLIni']=TL
                df_tableCIUni.at[unidad, 'TFLIni']=TLF

        # Ajustar tiempo fuera de linea
        elif vCommitU==0:
            TFL=0
            TL=0
            for i in range(24,0,-1):
                vCommitU=df_vComUni.loc[(df_vComUni.unidad==unidad)  & (df_vComUni.periodo==i),'valor'].iloc[0]
                if vCommitU==0:
                    TFL=TFL+1
                else:
                    df_tableCIUni.at[unidad, 'TFLIni']=TFL
                    df_tableCIUni.at[unidad, 'TLIni']=TL
                    break
            
            if i==1 and vCommitU==0:
                TFL=TFL+df_tableCIUni.at[unidad, 'TFLIni']
                df_tableCIUni.at[unidad, 'TFLIni']=TFL
                df_tableCIUni.at[unidad, 'TLIni']=TL
    
    df_tableCIUni=df_tableCIUni.reset_index()
    return df_tableCIUni
    


# In[14]:


def CalcularCIPlt(df_data,df_tableCIPlt):
    
    df_tableCIPlt=df_tableCIPlt.drop('fecha',axis=1).set_index('planta')

    df_VarGen= GetVariable(df_data,'vGenPlt',['planta','periodo'])
    df_VarGen['periodo']=df_VarGen['periodo'].astype(int)

    # df_vComPlt=GetVariable(df_data,'vCommitTrmPlt',['planta','periodo'])
    # df_vComPlt['periodo']=df_vComPlt['periodo'].astype(int)

    df_vConf= GetVariable(df_data,'vCommitTrmConf',['planta','periodo'])
    df_vConf['periodo']=df_vConf['periodo'].astype(int)

    df_vGenM2= GetVariable(df_data,'vGenTrmPltM2',['planta','periodo'])
    df_vGenM2['periodo']=df_vGenM2['periodo'].astype(int)

    df_vGenM12= GetVariable(df_data,'vGenTrmPltURM1',['planta','periodo'])
    df_vGenM12['periodo']=df_vGenM12['periodo'].astype(int)

    df_vGenM21= GetVariable(df_data,'vGenTrmPltDRM1',['planta','periodo'])
    df_vGenM21['periodo']=df_vGenM21['periodo'].astype(int)

    for planta in df_tableCIPlt.index:
        # unidad=df_tableCIPlt.loc[index,'unidad']
        # Ajustar generación
        Gen=df_VarGen.loc[(df_VarGen.planta==planta)  & (df_VarGen.periodo==24),'valor'].iloc[0]
        df_tableCIPlt.at[planta, 'GenIni']=round(Gen,0)

        conf=df_vConf.loc[(df_vConf.planta==planta)  & (df_vConf.periodo==24),'valor'].iloc[0]
        df_tableCIPlt.at[planta, 'ConfIni']=round(conf)

        GenM2=df_vGenM2.loc[(df_vGenM2.planta==planta)  & (df_vGenM2.periodo==24),'valor'].iloc[0]
        
        if GenM2>0:
            df_tableCIPlt.at[planta, 'TipoModIni']=2
        else:
            GenM2=df_vGenM12.loc[(df_vGenM12.planta==planta)  & (df_vGenM12.periodo==24),'valor'].iloc[0]
            if GenM2>0:
                df_tableCIPlt.at[planta, 'TipoModIni']=12 
            else:
                GenM2=df_vGenM21.loc[(df_vGenM21.planta==planta)  & (df_vGenM21.periodo==24),'valor'].iloc[0]
                if GenM2>0:
                    df_tableCIPlt.at[planta, 'TipoModIni']=21
                else:
                    df_tableCIPlt.at[planta, 'TipoModIni']=0                   

        # Ajustar tiempo en línea
        # vCommitU=df_vComPlt.loc[(df_vComPlt.planta==planta)  & (df_vComPlt.periodo==24),'valor'].iloc[0]
        # if vCommitU==1:
        #     TLF=0
        #     TL=0
        #     for i in range(24,0,-1):
        #         vCommitU=df_vComPlt.loc[(df_vComPlt.planta==planta)  & (df_vComPlt.periodo==i),'valor'].iloc[0]
        #         if vCommitU==1:
        #             TL=TL+1
        #         else:
        #             break

        #     TL=TL+df_tableCIPlt.at[planta, 'TLIni']
        #     df_tableCIPlt.at[planta, 'TLIni']=TL
        #     df_tableCIPlt.at[planta, 'TFLIni']=TLF

        # # Ajustar tiempo fuera de linea
        # elif vCommitU==0:
        #     TFL=0
        #     TL=0
        #     for i in range(24,0,-1):
        #         vCommitU=df_vComPlt.loc[(df_vComPlt.planta==planta)  & (df_vComPlt.periodo==i),'valor'].iloc[0]
        #         if vCommitU==0:
        #             TFL=TFL+1
        #         else:
        #             break

        #     TFL=TFL+df_tableCIPlt.at[planta, 'TFLIni']
        #     df_tableCIPlt.at[planta, 'TFLIni']=TFL
        #     df_tableCIPlt.at[planta, 'TLIni']=TL
    
    df_tableCIPlt=df_tableCIPlt.reset_index()

    return df_tableCIPlt


# In[15]:


def MmtosUnidad(df_tableDisp,sFecha,conn):

    # pTable='CIUni'
    # sFecha='2024-09-04'
    # Define your query
    query = "SELECT * FROM MmtosUnidad WHERE fechaIni<=cdate('" + sFecha + "') and fechaFin>=cdate('" + sFecha + "')"

    # Execute the query and store the result in a DataFrame
    df_table = pd.read_sql_query(query, conn) 
    df_table=df_table.set_index('unidad')

    for unidad in df_table.index:
        # if unidad=='PAIPA_4':
        #     stop=1
        #print(unidad)

        df_tableAux=df_table.loc[[unidad]]
        df_tableAux=df_tableAux.reset_index()

        for unidadAuxin in df_tableAux.index:

            fechaIni=df_tableAux.at[unidadAuxin,'fechaIni']
            fechaFin=df_tableAux.at[unidadAuxin,'fechaFin']
            if pd.to_datetime(fechaIni)==pd.to_datetime(sFecha):
                Pini=df_tableAux.at[unidadAuxin,'Pini']
            else:
                Pini=1

            if pd.to_datetime(fechaFin)==pd.to_datetime(sFecha):
                Pfin=df_tableAux.at[unidadAuxin,'Pfin']
            else:
                Pfin=24

            for per in range(Pini,Pfin+1):
                #print(per)
                df_tableDisp.loc[(df_tableDisp['unidad']==unidad) & (df_tableDisp['periodo']==per), 'valor'] = 0

    df_tableDisp = df_tableDisp.pivot(index='unidad', columns='periodo', values='valor')
    df_tableDisp = df_tableDisp.reset_index([])
    return df_tableDisp

def ShareThermalGen(dCENini,sSubarea,df_table,df_tablePesos):

    if dCENini>600:
        ni=4
    elif dCENini>300 and dCENini<=600:
        ni=3

    dCEN=round(dCENini/3,0)

    for i1 in range(1,ni):

        if i1==1:
            if sSubarea=='SubInterior':
                sUnidad='DC_TR_INT_Proy'
            elif sSubarea=='SubGCM':
                sUnidad='DC_TR_GCM_Proy'
            elif sSubarea=='SubCordoba-Sucre':
                sUnidad='DC_TR_CS_Proy'
            elif sSubarea=='SubAtlantico':
                sUnidad='DC_TR_ATL_Proy'
            elif sSubarea=='SubCerromatoso':
                sUnidad='DC_TR_CRR_Proy'
            elif sSubarea=='SubBolivar':
                sUnidad='DC_TR_BOL_Proy'
        elif i1==2:
            if sSubarea=='SubInterior':
                sUnidad='DC_TR_INT_Proy_1'
            elif sSubarea=='SubGCM':
                sUnidad='DC_TR_GCM_Proy_1'
            elif sSubarea=='SubCordoba-Sucre':
                sUnidad='DC_TR_CS_Proy_1'
            elif sSubarea=='SubAtlantico':
                sUnidad='DC_TR_ATL_Proy_1'
            elif sSubarea=='SubCerromatoso':
                sUnidad='DC_TR_CRR_Proy_1'
            elif sSubarea=='SubBolivar':
                sUnidad='DC_TR_BOL_Proy_1'
        elif i1==3:
            if sSubarea=='SubInterior':
                sUnidad='DC_TR_INT_Proy_2'
            elif sSubarea=='SubGCM':
                sUnidad='DC_TR_GCM_Proy_2'
            elif sSubarea=='SubCordoba-Sucre':
                sUnidad='DC_TR_CS_Proy_2'
            elif sSubarea=='SubAtlantico':
                sUnidad='DC_TR_ATL_Proy_2'
            elif sSubarea=='SubCerromatoso':
                sUnidad='DC_TR_CRR_Proy_2'
            elif sSubarea=='SubBolivar':
                sUnidad='DC_TR_BOL_Proy_2'

        for i in range(1,25):
            dvalor=round(dCEN,0)
            df_table.at[sUnidad,i] = round(df_table.at[sUnidad,i] + dvalor,2)

    
        if (sSubarea=='SubBolivar') or (sSubarea=='SubGCM') or (sSubarea=='SubAtlantico'):
            MaxVal = df_table.loc[sUnidad].max()
            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Caribe2'), 'peso'] = round((MaxVal * 0.4)/100,2)
            
            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Res_Car2'), 'peso'] = round(MaxVal/10,2)
            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Res_Car'), 'peso'] = round(MaxVal/10,2)
            if sSubarea=='SubBolivar':
                df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Bol'), 'peso'] = round((MaxVal * 0.4)/100,2)      
            if sSubarea=='SubAtlantico':
                df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Atl'), 'peso'] = round((MaxVal * 0.4)/100,2)
            if sSubarea=='SubGCM':
                df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_GCM'), 'peso'] = round((MaxVal * 0.08)/100,2)
                
        if (sSubarea=='SubCerromatoso') or (sSubarea=='SubCordoba-Sucre'):
            MaxVal = df_table.loc[sUnidad].max()
            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Res_Car'), 'peso'] = round(MaxVal/10,2)
            
    return 1

# In[16]:
def EvalProyectos(df_table,sFecha,conn):
    # Define your query
    query = "SELECT * FROM ProyGen WHERE fecha<=cdate('" + sFecha + "')"
    # Execute the query and store the result in a DataFrame
    df_Proy = pd.read_sql_query(query, conn) 

    query = "SELECT * FROM FactoresGen"
    # Execute the query and store the result in a DataFrame
    Factores = pd.read_sql_query(query, conn) 

    # Consultar tabla de pesos base
    df_tablePesos=ExecuteQuery(pTable='PesosZonas',sFecha='1900-12-01',conn=conn)

    if df_Proy.shape[0]>0:
        # Hacer agrupamiento por tipo de despacho, tipo de recurso y subarea
        df_grupo=df_Proy.groupby(['tipo_desp','tipo','subarea'])[['cen']].sum().reset_index()

        for ind in df_grupo.index:
            # Reset bandera para el recuros Térmica DC
            EjeShareThermalGen=0

            # Obtener los parámetros
            sTipoDesp=df_grupo.at[ind,'tipo_desp']
            sTipo=df_grupo.at[ind,'tipo']
            sSubarea=df_grupo.at[ind,'subarea']
            dCEN=df_grupo.at[ind,'cen']

            if sTipoDesp=='DC':
                if sTipo=='Eólico':
                    if sSubarea=='SubInterior':
                        sUnidad='EOLICO_INT_Proy'
                    elif sSubarea=='SubGCM':
                        sUnidad='EOLICO_GCM_Proy'
                    elif sSubarea=='SubCordoba-Sucre':
                        sUnidad='EOLICO_CS_Proy'
                    elif sSubarea=='SubAtlantico':
                        sUnidad='EOLICO_ATL_Proy'
                    elif sSubarea=='SubCerromatoso':
                        sUnidad='EOLICO_CRR_Proy'
                    elif sSubarea=='SubBolivar':
                        sUnidad='EOLICO_BOL_Proy'
                    else:
                        print("No se encontró la subarea ",sSubarea," por favor revisar")

                elif sTipo=='Solar':
                    if sSubarea=='SubInterior':
                        sUnidad='SOLAR_INT_Proy'
                    elif sSubarea=='SubGCM':
                        sUnidad='SOLAR_GCM_Proy'
                    elif sSubarea=='SubCordoba-Sucre':
                        sUnidad='SOLAR_CS_Proy'
                    elif sSubarea=='SubAtlantico':
                        sUnidad='SOLAR_ATL_Proy'
                    elif sSubarea=='SubCerromatoso':
                        sUnidad='SOLAR_CRR_Proy'
                    elif sSubarea=='SubBolivar':
                        sUnidad='SOLAR_BOL_Proy'
                    else:
                        print("No se encontró la subarea ",sSubarea," por favor revisar")

                elif sTipo=='Hidráulico':
                    if sSubarea=='SubInterior':
                        sUnidad='DC_HD_INT_Proy'
                    elif sSubarea=='SubGCM':
                        sUnidad='DC_HD_GCM_Proy'
                    elif sSubarea=='SubCordoba-Sucre':
                        sUnidad='DC_HD_CS_Proy'
                    elif sSubarea=='SubAtlantico':
                        sUnidad='DC_HD_ATL_Proy'
                    elif sSubarea=='SubCerromatoso':
                        sUnidad='DC_HD_CRR_Proy'
                    elif sSubarea=='SubBolivar':
                        sUnidad='DC_HD_BOL_Proy'
                    else:
                        print("No se encontró la subarea ",sSubarea," por favor revisar")

                elif sTipo=='Térmico':
                    
                    if dCEN>300:
                        # Procesar cuando la CEN es mayor que 300 para que se distribura en los diferentes segmentos
                        EjeShareThermalGen=ShareThermalGen(dCEN,sSubarea,df_table,df_tablePesos)
                        continue
                    else:
                        if sSubarea=='SubInterior':
                            sUnidad='DC_TR_INT_Proy'
                        elif sSubarea=='SubGCM':
                            sUnidad='DC_TR_GCM_Proy'
                        elif sSubarea=='SubCordoba-Sucre':
                            sUnidad='DC_TR_CS_Proy'
                        elif sSubarea=='SubAtlantico':
                            sUnidad='DC_TR_ATL_Proy'
                        elif sSubarea=='SubCerromatoso':
                            sUnidad='DC_TR_CRR_Proy'
                        elif sSubarea=='SubBolivar':
                            sUnidad='DC_TR_BOL_Proy'
                        else:
                            print("No se encontró la subarea ",sSubarea," por favor revisar")

                else:
                    print("No se encontró el tipo ",sTipo," por favor revisar")

            elif sTipoDesp=='NDC':
                if sSubarea=='SubInterior':
                    sUnidad='NDC_Proy'
                elif sSubarea=='SubGCM':
                    sUnidad='NDC_GCM_Proy'
                elif sSubarea=='SubCordoba-Sucre':
                    sUnidad='NDC_CS_Proy'
                elif sSubarea=='SubAtlantico':
                    sUnidad='NDC_ATL_Proy'
                elif sSubarea=='SubCerromatoso':
                    sUnidad='NDC_CRR_Proy'
                elif sSubarea=='SubBolivar':
                    sUnidad='NDC_BOL_Proy'
                else:
                    print("No se encontró la subarea ",sSubarea," por favor revisar")
            else:
                print("No se encuentra el tipo ", sTipoDesp, " por favor revisar")
                continue
            
            for i in range(1,25):
                # print(i,sTipoDesp,sTipo,sSubarea,dCEN)
                dFactor=Factores.loc[(Factores['tipo_desp'] == sTipoDesp) & (Factores['tipo_plt'] == sTipo) & (Factores['periodo'] == i), 'factor'].values[0]
                dvalor=round(dCEN*dFactor,2)
                df_table.at[sUnidad,i] = round(df_table.at[sUnidad,i] + dvalor,2)

            if sTipoDesp=='DC':
                if (sSubarea=='SubBolivar') or (sSubarea=='SubGCM') or (sSubarea=='SubAtlantico'):
                    MaxVal = df_table.loc[sUnidad].max()
                    if sTipo=='Eólico':
                        df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Caribe2'), 'peso'] = round((MaxVal * 0.12)/100,2)
                        
                        df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Res_Car2'), 'peso'] = round(MaxVal/10,2)
                        df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Res_Car'), 'peso'] = round(MaxVal/10,2)

                        if sSubarea=='SubBolivar':
                            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Bol'), 'peso'] = round((MaxVal * 0.07)/100,2)
                        if sSubarea=='SubAtlantico':
                            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Atl'), 'peso'] = round((MaxVal * 0.12)/100,2)
                        if sSubarea=='SubGCM':
                            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_GCM'), 'peso'] = round((MaxVal * 0.05)/100,2)

                                                
                    elif sTipo=='Solar':
                        df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Caribe2'), 'peso'] = round((MaxVal * 0.12)/100,2)
                        
                        df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Res_Car2'), 'peso'] = round(MaxVal/10,2)
                        df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Res_Car'), 'peso'] = round(MaxVal/10,2)
                        if sSubarea=='SubBolivar':
                            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Bol'), 'peso'] = round((MaxVal * 0.07)/100,2)
                        if sSubarea=='SubAtlantico':
                            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Atl'), 'peso'] = round((MaxVal * 0.12)/100,2)
                        if sSubarea=='SubGCM':
                            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_GCM'), 'peso'] = round((MaxVal * 0.05)/100,2)
                           
                    else:
                        df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Caribe2'), 'peso'] = round((MaxVal * 0.4)/100,2)
                        
                        df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Res_Car2'), 'peso'] = round(MaxVal/10,2)
                        df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Res_Car'), 'peso'] = round(MaxVal/10,2)
                        if sSubarea=='SubBolivar':
                            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Bol'), 'peso'] = round((MaxVal * 0.4)/100,2)      
                        if sSubarea=='SubAtlantico':
                            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Atl'), 'peso'] = round((MaxVal * 0.4)/100,2)
                        if sSubarea=='SubGCM':
                            df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_GCM'), 'peso'] = round((MaxVal * 0.08)/100,2)

                if (sSubarea=='SubCerromatoso') or (sSubarea=='SubCordoba-Sucre'):
                    MaxVal = df_table.loc[sUnidad].max()
                    df_tablePesos.loc[(df_tablePesos['unidad'] == sUnidad) & (df_tablePesos['zona'] == 'Z_Res_Car'), 'peso'] = round(MaxVal/10,2)

    return df_table, df_tablePesos

    


# In[17]:


def PrintFilesMod(d_Fecha,s_mainpath,conn,Horizon):
    
    # Archivo de disponibilidad unidad
    df_table=ExecuteQuery(pTable='DispUniBase',sFecha='1900-12-01',conn=conn)

    # Modificar la disponibilidad con los mantenimietnos
    df_table=MmtosUnidad(df_table,sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)

    # Para una ejecución de LP modificar la disponibilidad de acuerdo a la entrada de los proyectos
    if Horizon=='LT':
        [df_table,df_tablePesos]=EvalProyectos(df_table,sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)

        # Imprimir los pesos de las zonas de acuerdo a los proyectos que ingresan
        df_tablePesos = df_tablePesos.pivot(index='unidad', columns='zona', values='peso')
        df_tablePesos = df_tablePesos.reset_index()
        printFile(df_tablePesos,fileName='tmp_UnitWeights.txt',index='unidad',s_mainpath=s_mainpath)       

    df_table.reset_index(inplace=True)
    printFile(df_table,fileName='tmp_DispU.txt',index='unidad',s_mainpath=s_mainpath)

    # Archivo de precios por configuración
    df_table=ExecuteQuery(pTable='PreciosPltConf',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
    df_table['conf'] ='.c' + df_table['conf'].astype(str)
    df_table['Plt_Conf']=df_table['planta'].str.ljust(30) + '   ' + df_table['conf']
    df_table=df_table.drop(['planta','conf'],axis=1)
    printFile(df_table,fileName='tmp_OfferConfigThermalCC.txt',index='Plt_Conf',s_mainpath=s_mainpath)

    # Archivo de precios por planta hidráulica
    df_table=ExecuteQuery(pTable='PrecioPltH',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
    printFile(df_table,fileName='tmp_HydroPltPar.txt',index='planta',s_mainpath=s_mainpath)

    # Archivo de precios por planta térmica
    df_table=ExecuteQuery(pTable='PrecioPltTrm',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
    printFile(df_table,fileName='tmp_ThermalPltPar.txt',index='planta',s_mainpath=s_mainpath)

    # Archivo de zonas de seguridad unidades equivalentes
    df_table=ExecuteQuery(pTable='ZonaReqUE',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
    df_table = df_table.pivot(index='zona', columns='periodo', values='valor')
    df_table = df_table.reset_index()
    printFile(df_table,fileName='tmp_UnitZoneReq.txt',index='zona',s_mainpath=s_mainpath)

    # Archivo de zonas de seguridad MW mínimos
    df_table=ExecuteQuery(pTable='ZonaReqMin',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
    df_table = df_table.pivot(index='zona', columns='periodo', values='valor')
    df_table = df_table.reset_index()
    printFile(df_table,fileName='tmp_MWminZoneReq.txt',index='zona',s_mainpath=s_mainpath)

    # Archivo de zonas de seguridad MW máximos
    df_table=ExecuteQuery(pTable='ZonaReqMax',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
    df_table = df_table.pivot(index='zona', columns='periodo', values='valor')
    df_table = df_table.reset_index()
    printFile(df_table,fileName='tmp_MWmaxZoneReq.txt',index='zona',s_mainpath=s_mainpath)

    # Archivo de mínimos de generación
    df_table=ExecuteQuery(pTable='MinGen',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
    df_table = df_table.pivot(index='planta', columns='periodo', values='valor')
    df_table = df_table.reset_index()
    df_table.fillna(0,inplace=True)
    printFile(df_table,fileName='tmp_MinimumGenPlt.txt',index='planta',s_mainpath=s_mainpath)

    # Archivo de máximos de generación
    df_table=ExecuteQuery(pTable='MaxGen',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
    df_table = df_table.pivot(index='planta', columns='periodo', values='valor')
    df_table = df_table.reset_index()
    printFile(df_table,fileName='tmp_MaxGenPlt.txt',index='planta',s_mainpath=s_mainpath)

    # Archivo de máximos de generación
    df_table=ExecuteQuery(pTable='LimImpArea',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
    df_table = df_table.pivot(index='area', columns='periodo', values='valor')
    df_table = df_table.reset_index()
    printFile(df_table,fileName='tmp_AreaImportLim.txt',index='area',s_mainpath=s_mainpath)

    # Archivo de demanda
    df_table=ExecuteQuery(pTable='Demanda',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
    df_table['valor'] = df_table['valor'].round(4)
    df_table = df_table.pivot(index='periodo', columns='subarea', values='valor')
    df_table = df_table.reset_index()
    printFile(df_table,fileName='tmp_Demand.txt',index='periodo',s_mainpath=s_mainpath)

    # Archivo de parámetros
    df_table=ExecuteQuery(pTable='Parametros',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
    df_table['valor'] = df_table['valor'].round(4)
    printFile(df_table,fileName='tmp_Param2.txt',index='parametro',s_mainpath=s_mainpath)
    


# In[18]:


def PrintResults(df_data,s_mainpath,d_Fecha,cursor,conn,TipoSol,TipoMod):

    # Escribir en BD la generación de la planta
    PrintDataBD(df_data, pVar='vGenPlt', pIndex=['planta','periodo'], pTabla='GeneracionPlt', pCampos=['fecha','planta','periodo','valor']
                ,sfecha=d_Fecha.strftime('%Y-%m-%d'),pVarMod='Mod',cursor=cursor,conn=conn)

    # Escribir en BD la generación de la unidad
    PrintDataBD(df_data, pVar='vGenUni', pIndex=['unidad','periodo'], pTabla='GeneracionUni', pCampos=['fecha','unidad','periodo','valor']
                ,sfecha=d_Fecha.strftime('%Y-%m-%d'),pVarMod='Mod',cursor=cursor,conn=conn)
    
    # Escribir en BD la configuración de cada planta
    PrintDataBD(df_data, pVar='vCommitTrmConf', pIndex=['planta','periodo'], pTabla='ConfiguracionPlt', pCampos=['fecha','planta','periodo','valor']
                ,sfecha=d_Fecha.strftime('%Y-%m-%d'),pVarMod='Mod',cursor=cursor,conn=conn)

    if (TipoSol==0) & (TipoMod!='DESPACHO') :
        # Escribir en BD la configuración de cada planta
        PrintDataBD(df_data, pVar='vMarginal', pIndex=['periodo'], pTabla='TR_Marginal', pCampos=['fecha','periodo','valor']
                ,sfecha=d_Fecha.strftime('%Y-%m-%d'),pVarMod='Mod',cursor=cursor,conn=conn)
        

# In[19]:


def CIfromFile(s_mainpath):

    # Inicialización de las condiciones iniciales en base de datos
    file=r'Modules\InputData\tmp_ThermalPltCI_Init.txt'
    s_filepath=s_mainpath.joinpath(file)
    df_PltCI= pd.read_csv(s_filepath,delim_whitespace=True,encoding="ISO-8859-1")
    df_PltCI=df_PltCI[['ConfIni', 'GenIni','TipoModIni']]
    df_PltCI=df_PltCI.reset_index().rename(columns={'index': 'planta'})

    file=r'Modules\InputData\tmp_ThermalUniCI_Init.txt'
    s_filepath=s_mainpath.joinpath(file)
    df_UniCI= pd.read_csv(s_filepath,delim_whitespace=True,encoding="ISO-8859-1")
    df_UniCI=df_UniCI[['GenIni', 'DispIni', 'TDispIni', 'TLIni', 'TFLIni', 'CombIni']]
    df_UniCI=df_UniCI.reset_index().rename(columns={'index': 'unidad'})

    return df_PltCI,df_UniCI


# In[20]:


from openpyxl import load_workbook

def fun_CargarExcel(s_mainpath,file,cursor,conn,Horizon):

    # Cargar información a la base de datos a partir del archivo de excel
    # file=r'Data\DatosEntradaCOPS_' + sCaso + '_Abr25_Jul25.xlsx'
    s_filepath=str(s_mainpath.joinpath(file))
    # Load the Excel file
    # Load the workbook
    # with load_workbook(filename=s_filepath, read_only=True) as workbook:
    workbook=load_workbook(filename=s_filepath, read_only=True)
        # Get the sheet names
    sheet_names = workbook.sheetnames
        
    # Get all sheet names
    #print(f"Sheet names: {sheet_names}")

    if Horizon=='MT':
        l_sheets=['DispUniBase', 'DispConfPltBase', 'PreciosPltConf', 'PrecioPltTrm', 'PrecioPltH','MinGen','MaxGen','ZonaReqUE', 'ZonaReqMin', 
                  'ZonaReqMax', 'Demanda','LimImpArea','Parametros']
    elif Horizon=='ST':
        l_sheets=['PrecioPltTrm', 'PrecioPltH']
    elif Horizon=='LT':
        l_sheets=['DispUniBase', 'DispConfPltBase', 'PreciosPltConf', 'PrecioPltTrm', 'PrecioPltH','MinGen','MaxGen','ZonaReqUE', 'ZonaReqMin', 
                  'ZonaReqMax', 'Demanda','LimImpArea','Parametros','ProyGen','FactoresGen','PesosZonas']

    # Iterate over each sheet name
    for sheet in sheet_names:
        # Load sheet into a DataFrame
        if sheet in l_sheets:
            print(sheet)
            # Read Excel file with pandas
            df=pd.read_excel(s_filepath, sheet_name=sheet)
            if 'fecha' in df.columns:
                df['fecha']= pd.to_datetime(df['fecha'])

            if df.shape[0]>0:
                PrintDataBD(df, pVar='', pIndex=[], pTabla=sheet, pCampos=df.columns
                ,sfecha='1900-12-01',pVarMod='Entrada',cursor=cursor,conn=conn) 

    if 1==1:
        # Carga de la información de mmtos por unidad
        print('ManmtosUnidad')
        df_MmtosGen=pd.read_excel(s_filepath, sheet_name='ManmtosUnidad')
        df_MmtosGen['fechaIni']= pd.to_datetime(df_MmtosGen['fechaIni'])
        df_MmtosGen['fechaFin']= pd.to_datetime(df_MmtosGen['fechaFin'])
        PrintDataBD(df_MmtosGen, pVar='', pIndex=[], pTabla='MmtosUnidad', pCampos=df_MmtosGen.columns
        ,sfecha='1900-12-01',pVarMod='Entrada',cursor=cursor,conn=conn) 

    workbook.close()


# In[21]:


def ExecuteModel(s_mainpath,sCaso,fileDate,d_FechaIni,d_FechaFin,CIfile,Horizon,DR_LP,TipoSol,bMacroFiles,TipoMod):

    EndExecution=1

    # Ruta de la base de datos
    # Define the path to your Access database
    file=r'Modules\OutputData\ResultsGurobi_' + sCaso + '_' + fileDate + '.accdb'
    db_path = str(s_mainpath.joinpath(file))

    cursor,conn=ConexionBD(db_path)

    # Path to the GAMS executable (adjust based on your installation)
    gams_executable = r'C:\GAMS\win32\25.1\gams.exe'
    # Path to your GAMS file
    file=r'Main_COPS.gms'
    gams_file = s_mainpath.joinpath(file)

    #LP file
    file=r'Main_COPS.lp'
    LPfile=str(s_mainpath.joinpath(file))

    # Path to your batch (.bat) file
    bat_file_path = r'C:\Alejo\cops\ExecuteGams.bat'

    file_lp=r'Main_COPS.lp'
    file_lp=str(s_mainpath.joinpath(file_lp))


    # if os.path.exists('ModelStatus.txt'):
    #     os.remove('ModelStatus.txt')
    bEje=0
    ModelStatus=2
    NiterMax=0

    # Inicializar fecha
    delta=dt.timedelta(days=1)
    deltaNeg=dt.timedelta(days=-1)

    d_Fecha=d_FechaIni

    while d_Fecha<=d_FechaFin:

        # Read file to validate flag of execution or stop
        with open(r"C:\Alejo\cops\Modules\Utils\ArchivosAux\ExecutionFlag.txt", 'r', encoding='utf-8') as file:
            first_line = file.readline().strip()
        
        if int(first_line)==0:
            break


        print('################################################################################')
        print(d_Fecha.strftime('%Y-%m-%d'))
        print('################################################################################')

        if ((CIfile==1) and (d_Fecha==d_FechaIni)) or (ModelStatus==0):

            df_PltCI,df_UniCI=CIfromFile(s_mainpath)
            
            PrintDataBD(df_PltCI, pVar='', pIndex=[], pTabla='CIPlt', pCampos=['fecha','planta','ConfIni','GenIni','TipoModIni']
                        ,sfecha=d_Fecha.strftime('%Y-%m-%d'),pVarMod='Entrada',cursor=cursor,conn=conn)       
            PrintDataBD(df_UniCI, pVar='', pIndex=[], pTabla='CIUni', pCampos=['fecha','unidad','GenIni','DispIni','TDispIni','TLIni','TFLIni','CombIni']
                        ,sfecha=d_Fecha.strftime('%Y-%m-%d'),pVarMod='Entrada',cursor=cursor,conn=conn)     
        
        
        # Preproceso: Crear archivos para la ejecución de gams
        # Archivos condiciones iniciales unidad
        df_tableCIUni=ExecuteQuery(pTable='CIUni',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
        printFile(df_tableCIUni,fileName='tmp_ThermalUniCI.txt',index='unidad',s_mainpath=s_mainpath)

        # Archivos condiciones iniciales unidad
        df_tableCIPlt=ExecuteQuery(pTable='CIPlt',sFecha=d_Fecha.strftime('%Y-%m-%d'),conn=conn)
        printFile(df_tableCIPlt,fileName='tmp_ThermalPltCI.txt',index='planta',s_mainpath=s_mainpath)

        if bMacroFiles==0:
            # Imprimir archivos para el mdoelo
            PrintFilesMod(d_Fecha,s_mainpath,conn,Horizon=Horizon)

        if 1==1:
            # Run the GAMS file
            # result = subprocess.run([gams_executable, gams_file], capture_output=True, text=True)
            # # Print the output and error messages
            # print(result.stdout)
            # print(result.stderr)

            # Validar que si el archivo .lp cambia en cada itearación o ejecución
            # Capturar fecha inicial
            creation_timestamp = os.path.getmtime(file_lp)
            # Convert to human-readable format with seconds
            creation_datetime = dt.datetime.fromtimestamp(creation_timestamp)
            start_time = creation_datetime  # YYYY, MM, DD, HH, MM, SS
            print('Fecha creación del archivo lp',creation_datetime)


            # Run the batch file and wait for it to finish
            result = subprocess.run([bat_file_path], shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            print('Se creó el archivo lp de la fecha: ',d_Fecha.strftime('%Y-%m-%d'))

            # Capturar fecha final después de la creación con el proceso de gams
            creation_timestamp = os.path.getmtime(file_lp)
            # Convert to human-readable format with seconds
            creation_datetime = dt.datetime.fromtimestamp(creation_timestamp)
            end_time = creation_datetime

            print('Fecha Final',creation_datetime)

            # Compare
            if start_time >= end_time:
                print(f"No se está actualizando el archivo lp")
                model='No hay lp'
                break
            

            if TipoSol==1:
                # Ejecutar el Gurobi
                model=ExecuteGurobi(LPfile)

                if model.Status==2:
                    ModelStatus=2
                else:
                    ModelStatus=model.Status
            else:
                ModelStatus=2
                model=None
            
            
        if ModelStatus==2:
            NiterMax=0

            if TipoSol==1:
                #Obtener los datos del model Gurobi
                df_data=GetModeDataGurobi(model)
            else:
                df_data=GetModeDataGAMS(s_mainpath,TipoMod)

            # Imprimir resutlado
            PrintResults(df_data,s_mainpath,d_Fecha,cursor,conn,TipoSol,TipoMod)

            # Calcular las condiciones iniciales para el siguiente día
            # Para plantas
            
            iDay=d_Fecha.day
            if DR_LP==1 and iDay==5:
                d_FechaAux=d_Fecha + relativedelta(months=1)
                deltaD=dt.timedelta(days=-4)
                d_FechaAux=d_FechaAux + deltaD
            else:    
                d_FechaAux=d_Fecha+delta

            if TipoMod=="DESPACHO":
                df_tableCIPlt=CalcularCIPlt(df_data,df_tableCIPlt)
                df_tableCIPlt['fecha']=d_FechaAux.strftime('%Y-%m-%d')
                df_tableCIPlt['GenIni']=df_tableCIPlt['GenIni'].round(2)
                # df_tableCIPlt['ConfIni']=df_tableCIPlt['ConfIni'].round(0)
                PrintDataBD(df_tableCIPlt, pVar='', pIndex=[], pTabla='CIPlt', pCampos=['fecha','planta','ConfIni','GenIni','TipoModIni']
                            ,sfecha=d_FechaAux.strftime('%Y-%m-%d'),pVarMod='Entrada',cursor=cursor,conn=conn) 
                
                # df_tableCIUniIni=df_tableCIUni.copy()
                # Para unidades
                df_tableCIUni=CalcularCIUni(df_data,df_tableCIUni)
                df_tableCIUni['fecha']=d_FechaAux.strftime('%Y-%m-%d')
                df_tableCIUni['GenIni']=df_tableCIUni['GenIni'].round(2)
                PrintDataBD(df_tableCIUni, pVar='', pIndex=[], pTabla='CIUni', pCampos=['fecha','unidad','GenIni','DispIni','TDispIni','TLIni','TFLIni','CombIni']
                            ,sfecha=d_FechaAux.strftime('%Y-%m-%d'),pVarMod='Entrada',cursor=cursor,conn=conn) 
            

        else:
            ModelStatus=0
            with open('ModelStatus.txt', 'a') as file:  # 'a' mode opens the file for appending
                if bEje==0:
                    text_to_add = f"\n\n"
                    file.write(text_to_add)
                    text_to_add = f"*****************************\n"
                    file.write(text_to_add)                     
                    text_to_add = f"Inició ejecución {dt.datetime.now()}\n"
                    file.write(text_to_add)
                    bEje=1
                text_to_add = f"Estado en la fecha {pd.to_datetime(d_Fecha)} es {model.Status}\n"
                file.write(text_to_add)     

            

            # Condicional para terminar la ejecución si el modo finalizar está en 1 o si itera 3 veces en la misma fecha siendo infactible
            if (EndExecution==1) or (NiterMax==3):
                print(f"El proceso paró en la fecha: {d_Fecha.strftime('%Y-%m-%d')} con iteraciones {NiterMax}")
                break
            else:
                NiterMax=NiterMax+1
                d_Fecha=d_Fecha + deltaNeg

        iDay=d_Fecha.day
        if DR_LP==1 and iDay==5:
            d_Fecha = d_Fecha + relativedelta(months=1)
            deltaD=dt.timedelta(days=-4)
            d_Fecha=d_Fecha + deltaD
        else:    
            d_Fecha=d_Fecha + delta

    # Desconnectar la base de datos
    DesconexionDB(cursor,conn)

    return model


# ### Módulo principal para ejecución del modelo

# In[ ]:
if 1==1:
# try:

    # Ruta del archivo
    sFile=r"Parametros.json"
    script_dir = Path.cwd()
    script_dir = Path(__file__).resolve()
    script_dir=script_dir.parent.parent.parent
    sPathfile=os.path.join(script_dir,r"Modules\Utils\ArchivosAux",sFile)
    sPathfile

    # Open and load the JSON file
    with open(sPathfile,'r') as f:
        data = json.load(f)

    # Almancenar los parámetros en variables python
    Mod_Eje='GAE_COPS'


    year=data['Parametros']['Ano']
    mes=data['Parametros']['Mes']
    Carpeta=data['Parametros']['Carpeta']
    FechaInicial=data['Parametros']['Fecha_Inicial']
    FechaFinal=data['Parametros']['Fecha_Final']
    FileName=data['Parametros']['Nombre_Archivo']
    sCaso = data['Parametros']['Tipo_Caso']
    solver=data['Parametros']['Solver']
    TipoMod=data['Parametros']['TipoMod']
    Horizon=data['Parametros']['Tipo_Ejecucion']

    # Banderas de entrada al modelo
    bConsiderCI=data['BanderasMod']['CB_CI']
    bLoadExcel=data['BanderasMod']['LoadData']


    # In[ ]:


    # Transformar el formato d/m/y a y-m-d
    d_FechaIni = dt.datetime.strptime(FechaInicial, '%d/%m/%Y').strftime('%Y-%m-%d')
    d_FechaFin = dt.datetime.strptime(FechaFinal, '%d/%m/%Y').strftime('%Y-%m-%d')

    l_FechasIni=[]
    l_FechasIni.append(d_FechaIni)

    l_FechasFin=[]
    l_FechasFin.append(d_FechaFin)

    l_FileDate=[]
    l_FileDate.append(FileName)

    # Tipo de ejecución Gurobi o GAMS
    # 1 Gurobi
    # 0 GAMS
    if solver=='Gurobi':
        TipoSol=1
    else:
        TipoSol=0

    # Cargar las condiciones iniciales del archivo de texto generado por la macro
    if bConsiderCI=='Verdadero':
        CIfile=1
    else:
        CIfile=0

    # Cargar el archivos de Excel a la base de datos
    if bLoadExcel=='Verdadero':
        CargarExcel=1
    else:
        CargarExcel=0

    # Bandera para ejecución desde la macro
    if Horizon=='ST':
        bMacroFiles=1
        DR_LP=0
    elif Horizon=='MT':
        bMacroFiles=0
        DR_LP=0
    elif Horizon=='LT':
        bMacroFiles=0
        DR_LP=1
    else:
        print('Debe seleccionar un horizonte de tiempo')
        


    #Get main path and other folders
    s_mainpath=Path.cwd()
    s_mainpath = Path(__file__).resolve()
    s_mainpath=s_mainpath.parent.parent.parent

    # Contador o variable para entrar a las fechas requeridas, el 0 es para la primera fecha de la lista
    rFechaIni=0
    rFechaFin=5

    if Mod_Eje=='GAE_COPS':

        for rFecha in range(len(l_FechasIni)):

            if rFecha>=rFechaIni and rFecha<=rFechaFin:

                sFechaIni=l_FechasIni[rFecha]
                sFechaFin=l_FechasFin[rFecha]

                #sFechaFin='2025-03-31'

                #Camniar a formato fecha
                d_FechaIni=dt.datetime.strptime(sFechaIni,'%Y-%m-%d')
                d_FechaFin=dt.datetime.strptime(sFechaFin,'%Y-%m-%d')

                fileDate=l_FileDate[rFecha]

                Excelfile=r'Data\DatosEntradaCOPS_' + sCaso + '_' + fileDate + '.xlsx'

                # Ruta de la base de datos
                # Define the path to your Access database
                file=r'Modules\OutputData\ResultsGurobi_' + sCaso + '_' + fileDate + '.accdb'
                db_path = str(s_mainpath.joinpath(file))

                cursor, conn = ConexionBD(db_path)


                if CargarExcel==1:
                    fun_CargarExcel(s_mainpath,file=Excelfile,cursor=cursor,conn=conn,Horizon=Horizon)

                # Desconnectar la base de datos
                DesconexionDB(cursor,conn)

                model=ExecuteModel(s_mainpath,sCaso,fileDate,d_FechaIni,d_FechaFin,CIfile,Horizon,DR_LP,TipoSol,bMacroFiles,TipoMod)
    else:

        l_FechasIni=['2025-07-03']
        l_FechasFin=['2025-11-30']
        l_FileDate=['jul04_ago31']

        # l_FechasIni=['2025-09-23']
        # l_FechasFin=['2025-11-30']
        # l_FileDate=['jul04_ago31']

        # l_FechasIni=['2026-01-01','2026-05-01','2026-12-04','2027-01-01','2027-05-01','2027-10-10']
        # l_FechasFin=['2026-04-30','2026-08-31','2026-12-31','2027-04-30','2027-08-31','2027-12-31']
        # l_FileDate=['Ene26_Abr26','May26_Ago26','Sep26_Dic26','Ene27_Abr27','May27_Ago27','Sep27_Dic27']

        # l_FechasIni=['2035-01-01','2038-01-01','2041-01-01']
        # l_FechasFin=['2037-12-31','2040-12-31','2043-12-31']
        # l_FileDate=['LP_35_37','LP_38_40','LP_41_43']

        # l_FechasIni=['2029-01-01']
        # l_FechasFin=['2029-01-01']
        # l_FileDate=['LP_29_31']

        for rFecha in range(len(l_FechasIni)):

            if rFecha>=rFechaIni and rFecha<=rFechaFin:

                sFechaIni=l_FechasIni[rFecha]
                sFechaFin=l_FechasFin[rFecha]

                #sFechaFin='2025-03-31'

                #Camniar a formato fecha
                d_FechaIni=dt.datetime.strptime(sFechaIni,'%Y-%m-%d')
                d_FechaFin=dt.datetime.strptime(sFechaFin,'%Y-%m-%d')

                Nacional=1
                if Nacional==1:
                    sCaso='Nacional'
                else:
                    sCaso='Coordinado'

                fileDate=l_FileDate[rFecha]

                Excelfile=r'Data\DatosEntradaCOPS_' + sCaso + '_' + fileDate + '.xlsx'

                # Ruta de la base de datos
                # Define the path to your Access database
                file=r'Modules\OutputData\ResultsGurobi_' + sCaso + '_' + fileDate + '.accdb'
                db_path = str(s_mainpath.joinpath(file))

                cursor, conn = ConexionBD(db_path)


                # Banderas iniciales para la ejecución
                # Tipo de ejecución Gurobi o GAMS
                # 1 Gurobi
                # 0 GAMS
                TipoSol=1
                # Cargar las condiciones iniciales del archivo de texto generado por la macro
                CIfile=1
                # Ejecutar con archivos de la macro
                bMacroFiles=0
                # Cargar el archivos de Excel a la base de datos
                CargarExcel=1

                # Cargar la información de los proyectos de largo plazo
                # ST: Corto plazo para ejecutar un día y que no toma los valores del excel y no imprima de access
                # MT: Medio plazo
                # LT: Largo plazo
                Horizon='MT'
                
                # Bandera para pasar al siguiente mes en el día 5
                DR_LP=0

                if CargarExcel==1:
                    fun_CargarExcel(s_mainpath,file=Excelfile,cursor=cursor,conn=conn,Horizon=Horizon)


                # Desconnectar la base de datos
                DesconexionDB(cursor,conn)

                model=ExecuteModel(s_mainpath,sCaso,fileDate,d_FechaIni,d_FechaFin,CIfile,Horizon,DR_LP,TipoSol,bMacroFiles,TipoMod)

    print('El proceso terminó con éxito')
    # Crear ventana raíz oculta
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)  # ✅ Hace que los messagebox estén en primer plano
    messagebox.showinfo('Estado del proceso','Se el proceso terminó, validar la ventana de salida', parent=root)
    

# except:

#     root = tk.Tk()
#     root.withdraw()
#     root.attributes("-topmost", True)  # ✅ Hace que los messagebox estén en primer plano
#     messagebox.showerror('Estado del proceso','Error en el proceso, por favor validar', parent=root)   