#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from pathlib import Path
import os
import json
import datetime as dt
from ftplib import FTP_TLS
import tkinter as tk
from tkinter import messagebox
import csv
from openpyxl import load_workbook

root = tk.Tk()
root.withdraw()
root.attributes("-topmost", True)  # ✅ Hace que los messagebox estén en primer plano


# In[2]:


def PrintPrecio(file_path,tipo,Horizon):

    df_precioplt = pd.read_csv(file_path, sep=",", header=0)
    df_precioplt['Poferta']=(df_precioplt['Poferta']/1000).round(3)
    if tipo=='CS':
        df_precioplt=df_precioplt[(df_precioplt.TipoPlt=='CS')][['RecCops','Poferta']]
        df_precioplt = pd.concat([df_precioplt, pd.DataFrame({'RecCops': ['FLORES_I_CC'], 'Poferta': [10002]})], ignore_index=True)
        df_precioplt = pd.concat([df_precioplt, pd.DataFrame({'RecCops': ['FLORES_4_CC'], 'Poferta': [10003]})], ignore_index=True)
        df_precioplt = pd.concat([df_precioplt, pd.DataFrame({'RecCops': ['TEBSAB_CC'], 'Poferta': [10000]})], ignore_index=True)
        df_precioplt = pd.concat([df_precioplt, pd.DataFrame({'RecCops': ['TERMOCANDELARIA_CC'], 'Poferta': [10001]})], ignore_index=True)
        df_precioplt = pd.concat([df_precioplt, pd.DataFrame({'RecCops': ['TERMOCENTRO_CC'], 'Poferta': [10004]})], ignore_index=True)
        df_precioplt = pd.concat([df_precioplt, pd.DataFrame({'RecCops': ['TERMOEMCALI_CC'], 'Poferta': [10005]})], ignore_index=True)
        df_precioplt = pd.concat([df_precioplt, pd.DataFrame({'RecCops': ['TERMOSIERRA_CC'], 'Poferta': [10006]})], ignore_index=True)
        df_precioplt = pd.concat([df_precioplt, pd.DataFrame({'RecCops': ['TERMOVALLE_CC'], 'Poferta': [10007]})], ignore_index=True)
        SHEET_NAME = "PrecioPltTrm" 
    elif tipo=='H':
        df_precioplt=df_precioplt[(df_precioplt.TipoPlt=='H')][['RecCops','Poferta']]
        SHEET_NAME = "PrecioPltH" 

    df_precioplt = df_precioplt.sort_values(by='RecCops').reset_index(drop=True)

    if Horizon=='LT':
        file_path=rf'C:\Alejo\cops\Data\DatosEntradaCOPS_Base_LP.xlsx'
    else:    
        file_path=rf'C:\Alejo\cops\Data\DatosEntradaCOPS_Base.xlsx'
        
    # SHEET_NAME = "PrecioPltTrm"          # cámbialo
    TARGET_COL = "VarCost"        # a dónde escribir (columna de la hoja)

    # --- abrir libro y hoja ---
    wb  = load_workbook(file_path, data_only=True)
    ws  = wb [SHEET_NAME]

    wb_form = load_workbook(file_path)
    ws_form = wb_form[SHEET_NAME]


    # --- obtener fecha de A2 ---
    fecha_objetivo = ws["A2"].value  # puede ser datetime o string; comparamos por igualdad directa

    # --- localizar índices de columnas por encabezado ---
    # Buscamos encabezados en la fila 1
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    col_A = headers.get("fecha")
    col_planta = headers.get("planta")
    col_target = headers.get(TARGET_COL)

    if col_A is None or col_planta is None or col_target is None:
        raise ValueError("No se encontraron los encabezados 'fecha', 'planta' o el destino en la fila 1.")

    # --- construir mapa: planta -> fila (solo filas con la fecha de A2) ---
    mapa_filas = {}
    for r in range(2, 100):
        fecha_celda = ws.cell(row=r, column=col_A).value
        if fecha_celda == fecha_objetivo:
            planta = ws.cell(row=r, column=col_planta).value
            if planta is not None:
                key = str(planta)
                # si hay duplicados mismos (fecha, planta), te quedas con la última fila encontrada
                mapa_filas[key] = r

    # --- escribir valores y recolectar no encontrados ---
    # no_encontrados = []

    for rec, pof in df_precioplt[["RecCops", "Poferta"]].itertuples(index=False):
        key = str(rec)
        fila = mapa_filas.get(key)
        if fila is None:
            # no_encontrados.append(rec)
            messagebox.showinfo('Estado del proceso',f'No se encontró el precio de la planta {rec}', parent=root)
            continue
        # escribir el valor (convertimos a float por si llega como string)
        valor = float(pof) if pof is not None else None

        ws_form.cell(row=fila, column=col_target, value=valor)

    # --- guardar ---
    wb_form.save(file_path)


# In[3]:


def PrintPrecioCC(file_path,Horizon):

    df_precioplt = pd.read_csv(file_path, sep=",", header=0)
    df_precioplt=df_precioplt[['RecCops','Conf','Valor']]
    df_precioplt = df_precioplt.sort_values(by=['RecCops','Conf']).reset_index(drop=True)
    df_precioplt['Valor']=(df_precioplt['Valor']/1000).round(3)

    if Horizon=='LT':
        file_path=rf'C:\Alejo\cops\Data\DatosEntradaCOPS_Base_LP.xlsx'
    else:
        file_path=rf'C:\Alejo\cops\Data\DatosEntradaCOPS_Base.xlsx'

    # SHEET_NAME = "PrecioPltTrm"          # cámbialo
    TARGET_COL = "VarCost"        # a dónde escribir (columna de la hoja)
    SHEET_NAME = "PreciosPltConf" 

    # --- abrir libro y hoja ---
    wb  = load_workbook(file_path, data_only=True)
    ws  = wb [SHEET_NAME]

    wb_form = load_workbook(file_path)
    ws_form = wb_form[SHEET_NAME]

    # --- obtener fecha de A2 ---
    fecha_objetivo = ws["A2"].value  # puede ser datetime o string; comparamos por igualdad directa

    # --- localizar índices de columnas por encabezado ---
    # Buscamos encabezados en la fila 1
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    col_A = headers.get("fecha")
    col_planta = headers.get("planta")
    col_conf = headers.get("conf")
    col_target = headers.get(TARGET_COL)

    if col_A is None or col_planta is None or col_conf is None or col_target is None:
        raise ValueError("No se encontraron los encabezados 'fecha', 'planta' o el destino en la fila 1.")

    # --- construir mapa: planta -> fila (solo filas con la fecha de A2) ---
    mapa_filas = {}
    for r in range(2, 100):
        fecha_celda = ws.cell(row=r, column=col_A).value
        if fecha_celda == fecha_objetivo:
            planta = ws.cell(row=r, column=col_planta).value
            if planta is not None:
                conf = ws.cell(row=r, column=col_conf).value
                if conf is not None:
                    key = str(planta) + '.' + str(conf)
                    # si hay duplicados mismos (fecha, planta), te quedas con la última fila encontrada
                    mapa_filas[key] = r

    # --- escribir valores y recolectar no encontrados ---
    # no_encontrados = []

    for rec,conf, pof in df_precioplt[['RecCops','Conf','Valor']].itertuples(index=False):
        key = str(rec) + '.' + str(conf)
        fila = mapa_filas.get(key)
        if fila is None:
            # no_encontrados.append(rec)
            messagebox.showinfo('Estado del proceso',f'No se encontró el precio de la planta {rec}', parent=root)
            continue
        # escribir el valor (convertimos a float por si llega como string)
        valor = float(pof) if pof is not None else None

        ws_form.cell(row=fila, column=col_target, value=valor)

    # --- guardar ---
    wb_form.save(file_path)


# In[ ]:

# if 1==1:
try:


    # In[4]:


    sFile=r"Parametros.json"
    script_dir = Path.cwd()
    script_dir=script_dir.parent.parent
    sPathfile=os.path.join(script_dir,r"Modules\Utils\ArchivosAux",sFile)
    sPathfile = os.path.join(r"C:\Alejo\cops\Modules\Utils\ArchivosAux",sFile)

    # Open and load the JSON file
    with open(sPathfile,'r') as f:
        data = json.load(f)

    # Almancenar los parámetros en variables python
    year=data['Parametros']['Ano']
    mes=data['Parametros']['Mes']
    Carpeta=data['Parametros']['Carpeta']
    FechaInicial=data['Parametros']['Fecha_Inicial']
    FechaFinal=data['Parametros']['Fecha_Final']
    FileName=data['Parametros']['Nombre_Archivo']
    FechaDem=data['Parametros']['Fecha_Demanda']
    Horizon=data['Parametros']['Tipo_Ejecucion']

    sPathAE=data['Paths']['AEnerPath']


    # Obtener la fecha del día 1
    fecha_dt = dt.datetime.strptime(FechaInicial, "%d/%m/%Y")
    diaini = fecha_dt.day
    mesini=fecha_dt.month
    yearini=fecha_dt.year


    # In[16]:


    file_path=os.path.join(sPathAE + f"{int(year):00d}" + "-" + f"{int(mes):02d}" ,Carpeta , f"Precios_plantas_{yearini}_{mesini}_{diaini}.csv")
    PrintPrecio(file_path,tipo='CS',Horizon=Horizon)
    PrintPrecio(file_path,tipo='H',Horizon=Horizon)


    # In[6]:


    file_path=os.path.join(sPathAE + f"{int(year):00d}" + "-" + f"{int(mes):02d}" ,Carpeta , f"Precios_plantasCC_{yearini}_{mesini}_{diaini}.csv")
    PrintPrecioCC(file_path,Horizon=Horizon)


# In[ ]:


    messagebox.showinfo('Estado del proceso','Se cargaron los precios de las plantas de manera correcta', parent=root)

except:

    messagebox.showerror('Estado del proceso','Error en el proceso de cargue del precio de las plantas', parent=root)  

