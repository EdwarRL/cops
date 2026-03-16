#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from pathlib import Path
import os
import json
import datetime as dt
from ftplib import FTP_TLS
import tkinter as tk
from tkinter import messagebox
import csv
from openpyxl import load_workbook

root = tk.Tk()
root.withdraw()
root.attributes("-topmost", True)  # ✅ Hace que los messagebox estén en primer plano


# In[ ]:


def PrintPrecioTPL(file_path,Horizon):

    df_precioplt = pd.read_csv(file_path, sep=",", header=0)
    df_precioplt=df_precioplt[['planta','concepto','precio']]
    # df_precioplt = df_precioplt.sort_values(by=['RecCops','Conf']).reset_index(drop=True)
    df_precioplt['precio']=df_precioplt['precio'].round(3)
    df_precioplt['concepto'] = df_precioplt['concepto'].str.extract(r'(\d+)').astype('Int64')

    s_parentpath=Path('C:\Alejo\cops\Data')
    filepath=s_parentpath.joinpath(s_parentpath,'Mapeos.xlsx')
    df_pltt = pd.read_excel(filepath, sheet_name='NomRecursos')

    df_precioplt=df_precioplt.merge(df_pltt,left_on=['planta'],right_on=['RecOfe'],how='inner')   
    df_precioplt=df_precioplt[['RecCops','concepto','precio']]

    if Horizon=='LT':
        file_path=rf'C:\Alejo\cops\Data\DatosEntradaCOPS_Base_LP.xlsx'
    else:
        file_path=rf'C:\Alejo\cops\Data\DatosEntradaCOPS_Base.xlsx'

    # SHEET_NAME = "PrecioPltTrm"          # cámbialo
    TARGET_COL = "VarCost"        # a dónde escribir (columna de la hoja)
    SHEET_NAME = "PreciosPltConf" 

    # --- abrir libro y hoja ---
    wb  = load_workbook(file_path, data_only=True)
    ws  = wb [SHEET_NAME]

    wb_form = load_workbook(file_path)
    ws_form = wb_form[SHEET_NAME]

    # --- obtener fecha de A2 ---
    fecha_objetivo = ws["A2"].value  # puede ser datetime o string; comparamos por igualdad directa

    # --- localizar índices de columnas por encabezado ---
    # Buscamos encabezados en la fila 1
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    col_A = headers.get("fecha")
    col_planta = headers.get("planta")
    col_conf = headers.get("conf")
    col_target = headers.get(TARGET_COL)

    if col_A is None or col_planta is None or col_conf is None or col_target is None:
        raise ValueError("No se encontraron los encabezados 'fecha', 'planta' o el destino en la fila 1.")

    # --- construir mapa: planta -> fila (solo filas con la fecha de A2) ---
    mapa_filas = {}
    for r in range(2, 100):
        fecha_celda = ws.cell(row=r, column=col_A).value
        if fecha_celda == fecha_objetivo:
            planta = ws.cell(row=r, column=col_planta).value
            if planta is not None:
                conf = ws.cell(row=r, column=col_conf).value
                if conf is not None:
                    key = str(planta) + '.' + str(conf)
                    # si hay duplicados mismos (fecha, planta), te quedas con la última fila encontrada
                    mapa_filas[key] = r

    # --- escribir valores y recolectar no encontrados ---
    # no_encontrados = []

    for rec,conf, pof in df_precioplt[['RecCops','concepto','precio']].itertuples(index=False):
        if rec in ['BARRANQUILLA_3','BARRANQUILLA_4']:
            continue
        else:
            key = str(rec) + '.' + str(conf)
            fila = mapa_filas.get(key)
            if fila is None:
                # no_encontrados.append(rec)
                messagebox.showinfo('Estado del proceso',f'No se encontró el precio de la planta {rec}', parent=root)
                continue
            # escribir el valor (convertimos a float por si llega como string)
            valor = float(pof) if pof is not None else None

            ws_form.cell(row=fila, column=col_target, value=valor)

    SHEET_NAME = "PrecioPltTrm" 
    ws  = wb [SHEET_NAME]
    ws_form = wb_form[SHEET_NAME]

# --- obtener fecha de A2 ---
    fecha_objetivo = ws["A2"].value  # puede ser datetime o string; comparamos por igualdad directa

    # --- localizar índices de columnas por encabezado ---
    # Buscamos encabezados en la fila 1
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    col_A = headers.get("fecha")
    col_planta = headers.get("planta")
    col_target = headers.get(TARGET_COL)

    if col_A is None or col_planta is None or col_target is None:
        raise ValueError("No se encontraron los encabezados 'fecha', 'planta' o el destino en la fila 1.")

    # --- construir mapa: planta -> fila (solo filas con la fecha de A2) ---
    mapa_filas = {}
    for r in range(2, 10):
        fecha_celda = ws.cell(row=r, column=col_A).value
        if fecha_celda == fecha_objetivo:
            planta = ws.cell(row=r, column=col_planta).value
            if planta is not None:
                key = str(planta)
                # si hay duplicados mismos (fecha, planta), te quedas con la última fila encontrada
                mapa_filas[key] = r

    # --- escribir valores y recolectar no encontrados ---
    # no_encontrados = []

    for rec,conf, pof in df_precioplt[['RecCops','concepto','precio']].itertuples(index=False):
        if rec in ['TEBSAB_CC','TERMOCANDELARIA_CC']:
            continue
        else:
            key = str(rec)
            fila = mapa_filas.get(key)
            if fila is None:
                # no_encontrados.append(rec)
                messagebox.showinfo('Estado del proceso',f'No se encontró el precio de la planta {rec}', parent=root)
                continue
            # escribir el valor (convertimos a float por si llega como string)
            valor = float(pof) if pof is not None else None

            ws_form.cell(row=fila, column=col_target, value=valor)


    # --- guardar ---
    wb_form.save(file_path)


# In[ ]:

# if 1==1:
try:


# In[3]:


    sFile=r"Parametros.json"
    script_dir = Path.cwd()
    script_dir=script_dir.parent.parent
    sPathfile=os.path.join(script_dir,r"Modules\Utils\ArchivosAux",sFile)
    sPathfile = os.path.join(r"C:\Alejo\cops\Modules\Utils\ArchivosAux",sFile)

    # Open and load the JSON file
    with open(sPathfile,'r') as f:
        data = json.load(f)

    # Almancenar los parámetros en variables python
    year=data['Parametros']['Ano']
    mes=data['Parametros']['Mes']
    Carpeta=data['Parametros']['Carpeta']
    FechaInicial=data['Parametros']['Fecha_Inicial']
    FechaFinal=data['Parametros']['Fecha_Final']
    FileName=data['Parametros']['Nombre_Archivo']
    FechaDem=data['Parametros']['Fecha_Demanda']
    Horizon=data['Parametros']['Tipo_Ejecucion']

    sPathAE=data['Paths']['AEnerPath']



    # Obtener la fecha del día 1
    fecha_dt = dt.datetime.strptime(FechaInicial, "%d/%m/%Y")
    diaini = fecha_dt.day
    mesini=fecha_dt.month
    yearini=fecha_dt.year


    # In[13]:


    file_path=os.path.join(sPathAE + f"{int(year):00d}" + "-" + f"{int(mes):02d}" ,Carpeta , f"Precios_TPL_{yearini}_{mesini}_{diaini}.csv")
    PrintPrecioTPL(file_path,Horizon)


    messagebox.showinfo('Estado del proceso','Se cargaron los precios de TPL de manera correcta', parent=root)

except:

    messagebox.showerror('Estado del proceso','Error en el proceso de cargue de los precios de TPL', parent=root)  